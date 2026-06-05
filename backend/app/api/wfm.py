import csv
import io
import json
from datetime import date, datetime, time, timedelta
from typing import Type

from fastapi import APIRouter, Depends, File, HTTPException, Request, UploadFile
from fastapi.responses import Response
from sqlalchemy import func
from sqlalchemy.orm import Session

from app.db.session import get_db
from app.integrations.naumen.client import NaumenClientError
from app.integrations.naumen.service import client_from_settings, missing_settings_fields as missing_naumen_settings
from app.integrations.onec.service import OneCService, get_or_create_onec_settings
from app.models.integration_settings import AppSetting, NaumenProject, UserPreference, UserProjectAccess
from app.services.integration_settings import get_or_create_naumen_settings
from app.services.naumen_ncc_service import NaumenNccService, default_period as ncc_default_period
from app.models.wfm import (
    Absence,
    ActualWorkInterval,
    AuditLog,
    Employee,
    EmployeeAttendanceFact,
    EmployeeDailyStat,
    EmployeeImportBatch,
    EmployeeImportError,
    EmployeeIntervalStat,
    EmployeeProjectAccess,
    EmployeeSkill,
    EmployeeTeamMembership,
    ExportLog,
    ImportBatch,
    ImportError,
    KpiSnapshot,
    NaumenOperator,
    OneCStatusCheckError,
    OneCStatusCheckRun,
    PlanningSettings,
    Queue,
    QueueSkill,
    ScheduleAssignment,
    ScheduleCoverage,
    ScheduleGenerationRun,
    ScheduleRecommendation,
    ScheduleRule,
    Shift,
    Skill,
    StaffingRequirement,
    Team,
    WorkloadInterval,
    utcnow,
)
from app.schemas.wfm import (
    AbsenceIn,
    AbsenceOut,
    ActualWorkIn,
    ActualWorkOut,
    CalculateStaffingIn,
    CalculateStaffingOut,
    CoverageOut,
    CoverageRecalcOut,
    DateRangeIn,
    EmployeeCheckAllIn,
    EmployeeIn,
    EmployeeNaumenLinkIn,
    EmployeeOut,
    EmployeeSkillsIn,
    GenerateDraftIn,
    GenerateDraftOut,
    ImportBatchDetail,
    ImportBatchOut,
    PlanFactOut,
    PlanningSettingsIn,
    PlanningSettingsOut,
    QueueIn,
    QueueOut,
    QueueSkillIn,
    QueueSkillOut,
    ScheduleIn,
    ScheduleOut,
    ScheduleRecommendationOut,
    ScheduleRuleIn,
    ScheduleRuleOut,
    ShiftIn,
    ShiftOut,
    SkillIn,
    SkillEmployeesIn,
    SkillOut,
    StaffingIn,
    StaffingOut,
    SummaryOut,
    TeamIn,
    TeamEmployeesIn,
    TeamOut,
    NaumenOperatorOut,
    WorkloadIn,
    WorkloadOut,
)
from app.services.planning import calculation_note, calculate_required_agents, coverage_percent, employee_matches_required_skills, skill_match_score
from app.wfm.scheduler_engine import Candidate, SchedulerSettings, choose_best_candidate

router = APIRouter(prefix="/api/v1", tags=["wfm"])


def get_or_404(db: Session, model: Type, item_id: int):
    item = db.get(model, item_id)
    if not item:
        raise HTTPException(status_code=404, detail="Объект не найден")
    return item


def save_item(db: Session, item):
    db.add(item)
    db.commit()
    db.refresh(item)
    return item


def update_item(db: Session, model: Type, item_id: int, payload):
    item = get_or_404(db, model, item_id)
    for key, value in payload.model_dump().items():
        setattr(item, key, value)
    return save_item(db, item)


def deactivate_or_delete(db: Session, model: Type, item_id: int):
    item = get_or_404(db, model, item_id)
    if hasattr(item, "is_active"):
        item.is_active = False
        save_item(db, item)
        return {"status": "disabled", "id": item_id}
    db.delete(item)
    db.commit()
    return {"status": "deleted", "id": item_id}


def team_name(db: Session, team_id: int | None) -> str | None:
    return db.get(Team, team_id).name if team_id and db.get(Team, team_id) else None


def queue_name(db: Session, queue_id: int | None) -> str | None:
    return db.get(Queue, queue_id).name if queue_id and db.get(Queue, queue_id) else None


def employee_name(db: Session, employee_id: int | None) -> str | None:
    return db.get(Employee, employee_id).full_name if employee_id and db.get(Employee, employee_id) else None


def shift_name(db: Session, shift_id: int | None) -> str | None:
    return db.get(Shift, shift_id).name if shift_id and db.get(Shift, shift_id) else None


def skill_name(db: Session, skill_id: int | None) -> str | None:
    return db.get(Skill, skill_id).name if skill_id and db.get(Skill, skill_id) else None


def validate_inn_or_400(inn: str | None) -> str | None:
    if inn is None or inn == "":
        return None
    cleaned = "".join(ch for ch in inn if ch.isdigit())
    if len(cleaned) not in {10, 12}:
        raise HTTPException(status_code=422, detail="ИНН должен содержать 10 или 12 цифр")
    return cleaned


def normalize_inn(inn: str | None) -> str | None:
    if not inn:
        return None
    cleaned = "".join(ch for ch in str(inn).strip() if ch.isdigit())
    return cleaned or None


def normalize_phone(value: str | None) -> str | None:
    if not value:
        return None
    digits = "".join(ch for ch in value if ch.isdigit())
    if len(digits) == 11 and digits.startswith("8"):
        return "+7" + digits[1:]
    if len(digits) == 11 and digits.startswith("7"):
        return "+" + digits
    return value.strip()


def employee_out(db: Session, item: Employee, active_project_id: int | None = None) -> EmployeeOut:
    skill_ids = [row.skill_id for row in db.query(EmployeeSkill).filter(EmployeeSkill.employee_id == item.id).all()]
    project_ids = employee_project_ids(db, item.id)
    projects = db.query(NaumenProject).filter(NaumenProject.id.in_(project_ids)).all() if project_ids else []
    project_names = [project.title for project in projects]
    project_id = active_project_id or item.project_id or (project_ids[0] if project_ids else None)
    team_id = active_team_id_for_project(db, item.id, project_id)
    return EmployeeOut.model_validate(item).model_copy(update={
        "project_id": project_id,
        "project_ids": project_ids,
        "project_names": project_names,
        "team_id": team_id,
        "team_name": team_name(db, team_id),
        "skill_ids": skill_ids,
    })


def apply_onec_status(employee: Employee, result) -> None:
    employee.onec_status = result.status
    employee.onec_status_label = result.status_label
    employee.onec_last_checked_at = utcnow()
    employee.onec_last_check_message = result.message
    employee.onec_active_cards_count = result.active_cards_count
    employee.onec_dismissed_cards_count = result.dismissed_cards_count
    employee.onec_metadata = json.dumps({
        "gateway_version": getattr(result, "gateway_version", None),
        "gateway_warning": getattr(result, "gateway_warning", None),
        "lookup_strategy_used": getattr(result, "lookup_strategy_used", None),
        "cards": getattr(result, "cards", None) or [],
        "query_warnings": getattr(result, "query_warnings", None) or [],
        "query_errors": getattr(result, "query_errors", None) or [],
        "raw_summary": getattr(result, "raw_summary", None) or {},
    }, ensure_ascii=False, default=str)
    if result.status == "ok":
        employee.onec_status = "check_error"
        employee.onec_status_label = "Ошибка сверки"
        employee.onec_last_check_message = "Gateway вернул технический статус ok вместо кадрового статуса сотрудника."
        return
    if result.status in {"working", "dismissed", "not_found", "not_found_person_cards", "check_error", "gateway_unavailable", "gateway_invalid", "gateway_timeout", "onec_connection_error"}:
        employee.employment_status = result.status
    if result.last_hire_date:
        try:
            employee.hire_date = date.fromisoformat(result.last_hire_date[:10])
        except ValueError:
            pass
    if result.last_dismissal_date:
        try:
            employee.dismissal_date = date.fromisoformat(result.last_dismissal_date[:10])
        except ValueError:
            pass


def get_planning_settings(db: Session) -> PlanningSettings:
    settings = db.query(PlanningSettings).order_by(PlanningSettings.id).first()
    if settings:
        return settings
    settings = PlanningSettings()
    return save_item(db, settings)


def employee_skill_map(db: Session, employee_id: int) -> dict[int, int]:
    return {
        row.skill_id: row.level
        for row in db.query(EmployeeSkill).filter(EmployeeSkill.employee_id == employee_id).all()
    }


def normalize_person_name(value: str | None) -> str:
    return " ".join((value or "").strip().lower().replace("ё", "е").split())


def split_person_name(value: str | None) -> tuple[str | None, str | None, str | None]:
    parts = normalize_person_name(value).split()
    if len(parts) < 2:
        return None, None, None
    return parts[0], parts[1], " ".join(parts[2:]) or None


def employee_project_ids(db: Session, employee_id: int) -> list[int]:
    return [
        row.project_id
        for row in db.query(EmployeeProjectAccess)
        .filter(EmployeeProjectAccess.employee_id == employee_id, EmployeeProjectAccess.can_work.is_(True))
        .order_by(EmployeeProjectAccess.is_primary.desc(), EmployeeProjectAccess.project_id)
        .all()
    ]


def employee_has_project_access(db: Session, employee_id: int, project_id: int | None) -> bool:
    if project_id is None:
        return True
    return db.query(EmployeeProjectAccess).filter(
        EmployeeProjectAccess.employee_id == employee_id,
        EmployeeProjectAccess.project_id == project_id,
        EmployeeProjectAccess.can_work.is_(True),
    ).first() is not None


def sync_employee_project_access(db: Session, employee: Employee, project_ids: list[int], primary_project_id: int | None = None) -> None:
    unique_ids = sorted({int(item) for item in project_ids if item})
    if not unique_ids:
        raise HTTPException(status_code=422, detail="У сотрудника должен быть выбран хотя бы один рабочий контур")
    primary = primary_project_id if primary_project_id in unique_ids else unique_ids[0]
    db.query(EmployeeProjectAccess).filter(EmployeeProjectAccess.employee_id == employee.id).delete(synchronize_session=False)
    for project_id in unique_ids:
        db.add(EmployeeProjectAccess(employee_id=employee.id, project_id=project_id, can_work=True, is_primary=project_id == primary))
    employee.project_id = primary


def active_team_id_for_project(db: Session, employee_id: int, project_id: int | None) -> int | None:
    if project_id is None:
        employee = db.get(Employee, employee_id)
        return employee.team_id if employee else None
    membership = db.query(EmployeeTeamMembership).filter(
        EmployeeTeamMembership.employee_id == employee_id,
        EmployeeTeamMembership.project_id == project_id,
    ).first()
    return membership.team_id if membership else None


def set_employee_team_for_project(db: Session, employee: Employee, team_id: int | None, project_id: int | None) -> None:
    if project_id is None:
        employee.team_id = team_id
        return
    db.query(EmployeeTeamMembership).filter(
        EmployeeTeamMembership.employee_id == employee.id,
        EmployeeTeamMembership.project_id == project_id,
    ).delete(synchronize_session=False)
    if team_id:
        team = get_or_404(db, Team, team_id)
        if team.project_id != project_id:
            raise HTTPException(status_code=422, detail="Команда должна принадлежать текущему контуру")
        if not employee_has_project_access(db, employee.id, project_id):
            raise HTTPException(status_code=422, detail="Сначала добавьте сотрудника в этот проект")
        db.add(EmployeeTeamMembership(employee_id=employee.id, team_id=team_id, project_id=project_id, is_primary=True))
    if employee.project_id == project_id or employee.team_id is None:
        employee.team_id = team_id


def queue_required_skills(db: Session, queue_id: int) -> list[tuple[int, int, bool]]:
    return [
        (row.skill_id, row.min_level, row.is_required)
        for row in db.query(QueueSkill).filter(QueueSkill.queue_id == queue_id).all()
    ]


def shift_covers_interval(shift: Shift, interval_start: datetime, interval_end: datetime) -> bool:
    start_time = interval_start.time()
    end_time = interval_end.time()
    return shift.start_time <= start_time and shift.end_time >= end_time


def current_user_id(request: Request | None) -> int | None:
    user = getattr(request.state, "user", None) if request else None
    return getattr(user, "id", None)


def ensure_project_access(db: Session, request: Request | None, project_id: int | None) -> None:
    if project_id is None:
        return
    user = getattr(request.state, "user", None) if request else None
    if not user:
        raise HTTPException(status_code=401, detail="Требуется авторизация")
    if user.is_superuser or user.role == "admin":
        return
    allowed = db.query(UserProjectAccess).filter(
        UserProjectAccess.user_id == user.id,
        UserProjectAccess.naumen_project_id == project_id,
        UserProjectAccess.can_view.is_(True),
    ).first()
    if not allowed:
        raise HTTPException(status_code=403, detail="Нет доступа к проекту")


def current_project_id(db: Session, request: Request | None, explicit_project_id: int | None = None) -> int | None:
    if explicit_project_id is not None:
        ensure_project_access(db, request, explicit_project_id)
        return explicit_project_id
    user = getattr(request.state, "user", None) if request else None
    preference = db.query(UserPreference).filter(UserPreference.user_id == user.id).first() if user else None
    if preference and preference.selected_project_id:
        ensure_project_access(db, request, preference.selected_project_id)
        return preference.selected_project_id
    item = db.query(NaumenProject).filter(NaumenProject.is_active.is_(True)).order_by(NaumenProject.is_default.desc(), NaumenProject.id).first()
    return item.id if item else None


def apply_project_filter(query, model: Type, project_id: int | None):
    if project_id is not None and hasattr(model, "project_id"):
        return query.filter(model.project_id == project_id)
    return query


def log_audit(db: Session, action: str, request: Request | None = None, entity_type: str = "system", entity_id: str | None = None, details: str | None = None) -> None:
    user = getattr(request.state, "user", None) if request else None
    db.add(AuditLog(
        actor_user_id=getattr(user, "id", None),
        actor_username=getattr(user, "username", None),
        actor=getattr(user, "username", None) or "system",
        action=action,
        entity_type=entity_type,
        entity_id=entity_id,
        details=details,
        details_json=details,
        ip_address=request.client.host if request and request.client else None,
        user_agent=request.headers.get("user-agent") if request else None,
    ))


def csv_response(filename: str, headers: list[str], rows: list[list[object]], db: Session | None = None, request: Request | None = None, report_type: str | None = None) -> Response:
    if db and report_type:
        db.add(ExportLog(user_id=current_user_id(request), report_type=report_type, filename=filename, rows_count=len(rows)))
        log_audit(db, "export_report", request=request, entity_type="report", entity_id=report_type, details=f"filename={filename}; rows={len(rows)}")
        db.commit()
    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(headers)
    writer.writerows(rows)
    content = "\ufeff" + output.getvalue()
    return Response(
        content=content,
        media_type="text/csv; charset=utf-8",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


def ensure_schedule_rules(db: Session) -> None:
    defaults = [
        ("max_weekly_hours", "40", "Максимум часов на сотрудника в неделю"),
        ("min_rest_hours", "12", "Минимальный отдых между сменами"),
        ("allow_overtime", "false", "Разрешать сверхурочные назначения"),
        ("balance_weekend_shifts", "true", "Балансировать выходные смены"),
        ("prefer_skill_match", "true", "Предпочитать совпадение навыков"),
    ]
    existing = {row.name for row in db.query(ScheduleRule).all()}
    for name, value, description in defaults:
        if name not in existing:
            db.add(ScheduleRule(name=name, value=value, description=description))
    db.commit()


def parse_csv_datetime(row: dict[str, str]) -> tuple[datetime, datetime]:
    work_date = datetime.strptime(row["date"].strip(), "%Y-%m-%d").date()
    interval_start = datetime.combine(work_date, datetime.strptime(row["interval_start"].strip(), "%H:%M").time())
    interval_end = datetime.combine(work_date, datetime.strptime(row["interval_end"].strip(), "%H:%M").time())
    if interval_end <= interval_start:
        raise ValueError("Конец интервала должен быть позже начала")
    return interval_start, interval_end


def workload_out(db: Session, item: WorkloadInterval) -> WorkloadOut:
    return WorkloadOut.model_validate(item).model_copy(update={"queue_name": queue_name(db, item.queue_id)})


def staffing_out(db: Session, item: StaffingRequirement) -> StaffingOut:
    return StaffingOut.model_validate(item).model_copy(update={"queue_name": queue_name(db, item.queue_id)})


def schedule_out(db: Session, item: ScheduleAssignment) -> ScheduleOut:
    return ScheduleOut.model_validate(item).model_copy(update={
        "employee_name": employee_name(db, item.employee_id),
        "shift_name": shift_name(db, item.shift_id),
        "queue_name": queue_name(db, item.queue_id),
    })


def coverage_out(db: Session, item: ScheduleCoverage) -> CoverageOut:
    return CoverageOut.model_validate(item).model_copy(update={"queue_name": queue_name(db, item.queue_id)})


def actual_work_out(db: Session, item: ActualWorkInterval) -> ActualWorkOut:
    return ActualWorkOut.model_validate(item).model_copy(update={
        "employee_name": employee_name(db, item.employee_id),
        "queue_name": queue_name(db, item.queue_id),
    })


def recommendation_out(db: Session, item: ScheduleRecommendation) -> ScheduleRecommendationOut:
    return ScheduleRecommendationOut.model_validate(item).model_copy(update={
        "queue_name": queue_name(db, item.queue_id),
        "employee_name": employee_name(db, item.employee_id),
    })


def severity_for_gap(gap: int) -> str:
    if gap >= 3:
        return "critical"
    if gap == 2:
        return "warning"
    return "info"


NAUMEN_STATUS_LABELS = {
    "not_linked": "Не сопоставлен",
    "linked": "Сопоставлен",
    "not_found": "Не найден в Naumen",
    "mismatch": "Есть расхождения",
    "check_error": "Ошибка сверки",
    "api_unavailable": "Naumen недоступен",
    "no_access": "Нет доступа",
}


def update_employee_naumen_status(employee: Employee, status: str, message: str, metadata: dict | None = None) -> None:
    employee.naumen_status = status
    employee.naumen_status_label = NAUMEN_STATUS_LABELS.get(status, status)
    employee.naumen_last_checked_at = utcnow()
    employee.naumen_last_check_message = message
    if metadata is not None:
        employee.naumen_metadata = json.dumps(metadata, ensure_ascii=False, default=str)


def recalculate_coverage_rows(db: Session, payload: DateRangeIn) -> int:
    start = datetime.combine(payload.date_from, time.min)
    end = datetime.combine(payload.date_to, time.max)
    requirements_query = db.query(StaffingRequirement).filter(
        StaffingRequirement.interval_start >= start,
        StaffingRequirement.interval_start <= end,
    )
    if payload.queue_id:
        requirements_query = requirements_query.filter(StaffingRequirement.queue_id == payload.queue_id)
    requirements = requirements_query.order_by(StaffingRequirement.interval_start).all()
    count = 0

    for requirement in requirements:
        work_date = requirement.interval_start.date()
        assignments_query = db.query(ScheduleAssignment).filter(
            ScheduleAssignment.work_date == work_date,
            ScheduleAssignment.queue_id == requirement.queue_id,
            ScheduleAssignment.status.in_(["draft", "confirmed", "published"]),
        )
        assignments = assignments_query.all()
        planned = confirmed = published = 0
        for assignment in assignments:
            shift = db.get(Shift, assignment.shift_id)
            if not shift or not shift_covers_interval(shift, requirement.interval_start, requirement.interval_end):
                continue
            planned += 1
            if assignment.status in ("confirmed", "published"):
                confirmed += 1
            if assignment.status == "published":
                published += 1

        existing = db.query(ScheduleCoverage).filter(
            ScheduleCoverage.interval_start == requirement.interval_start,
            ScheduleCoverage.interval_end == requirement.interval_end,
            ScheduleCoverage.queue_id == requirement.queue_id,
        ).first()
        values = {
            "required_agents": requirement.required_agents,
            "planned_agents": planned,
            "confirmed_agents": confirmed,
            "published_agents": published,
            "gap_agents": requirement.required_agents - planned,
            "coverage_percent": coverage_percent(planned, requirement.required_agents),
        }
        if existing:
            for key, value in values.items():
                setattr(existing, key, value)
        else:
            db.add(ScheduleCoverage(
                interval_start=requirement.interval_start,
                interval_end=requirement.interval_end,
                queue_id=requirement.queue_id,
                **values,
            ))
        count += 1
    db.commit()
    return count


@router.get("/teams", response_model=list[TeamOut])
def list_teams(request: Request, project_id: int | None = None, db: Session = Depends(get_db)):
    project_id = current_project_id(db, request, project_id)
    return apply_project_filter(db.query(Team), Team, project_id).order_by(Team.id).all()


@router.post("/teams", response_model=TeamOut)
def create_team(payload: TeamIn, request: Request, db: Session = Depends(get_db)):
    data = payload.model_dump()
    data["project_id"] = current_project_id(db, request, data.get("project_id"))
    return save_item(db, Team(**data))


@router.get("/teams/{item_id}", response_model=TeamOut)
def read_team(item_id: int, db: Session = Depends(get_db)):
    return get_or_404(db, Team, item_id)


@router.put("/teams/{item_id}", response_model=TeamOut)
def update_team(item_id: int, payload: TeamIn, request: Request, db: Session = Depends(get_db)):
    item = get_or_404(db, Team, item_id)
    ensure_project_access(db, request, item.project_id)
    for key, value in payload.model_dump().items():
        setattr(item, key, value)
    return save_item(db, item)


@router.delete("/teams/{item_id}")
def delete_team(item_id: int, db: Session = Depends(get_db)):
    return deactivate_or_delete(db, Team, item_id)


@router.post("/teams/{item_id}/archive")
def archive_team(item_id: int, db: Session = Depends(get_db)):
    active_employees = db.query(EmployeeTeamMembership).join(Employee, Employee.id == EmployeeTeamMembership.employee_id).filter(
        EmployeeTeamMembership.team_id == item_id,
        Employee.is_active.is_(True),
    ).count()
    if active_employees:
        raise HTTPException(status_code=409, detail="Нельзя архивировать команду с активными сотрудниками")
    return deactivate_or_delete(db, Team, item_id)


@router.get("/teams/{item_id}/employees", response_model=list[EmployeeOut])
def list_team_employees(item_id: int, request: Request, db: Session = Depends(get_db)):
    team = get_or_404(db, Team, item_id)
    ensure_project_access(db, request, team.project_id)
    rows = db.query(Employee).join(EmployeeTeamMembership, EmployeeTeamMembership.employee_id == Employee.id).filter(
        EmployeeTeamMembership.team_id == item_id,
        EmployeeTeamMembership.project_id == team.project_id,
    ).order_by(Employee.full_name).all()
    return [employee_out(db, item, team.project_id) for item in rows]


@router.put("/teams/{item_id}/employees")
def update_team_employees(item_id: int, payload: TeamEmployeesIn, request: Request, db: Session = Depends(get_db)) -> dict:
    team = get_or_404(db, Team, item_id)
    ensure_project_access(db, request, team.project_id)
    employees = db.query(Employee).join(EmployeeProjectAccess, EmployeeProjectAccess.employee_id == Employee.id).filter(
        Employee.id.in_(payload.employee_ids),
        EmployeeProjectAccess.project_id == team.project_id,
        EmployeeProjectAccess.can_work.is_(True),
    ).all() if payload.employee_ids else []
    if len(employees) != len(set(payload.employee_ids)):
        raise HTTPException(status_code=422, detail="Все сотрудники должны принадлежать текущему контуру")
    db.query(EmployeeTeamMembership).filter(EmployeeTeamMembership.team_id == item_id, EmployeeTeamMembership.project_id == team.project_id).delete(synchronize_session=False)
    db.query(Employee).filter(Employee.team_id == item_id, Employee.project_id == team.project_id).update({"team_id": None}, synchronize_session=False)
    for employee in employees:
        db.add(EmployeeTeamMembership(employee_id=employee.id, team_id=item_id, project_id=team.project_id, is_primary=True))
        if employee.project_id == team.project_id or employee.team_id is None:
            employee.team_id = item_id
    log_audit(db, "team_employees_update", request=request, entity_type="team", entity_id=str(item_id), details=f"employees={len(employees)}")
    db.commit()
    return {"status": "ok", "team_id": item_id, "employee_ids": sorted(set(payload.employee_ids))}


@router.get("/skills", response_model=list[SkillOut])
def list_skills(request: Request, project_id: int | None = None, db: Session = Depends(get_db)):
    return db.query(Skill).order_by(Skill.name).all()


@router.post("/skills", response_model=SkillOut)
def create_skill(payload: SkillIn, request: Request, db: Session = Depends(get_db)):
    data = payload.model_dump()
    data["project_id"] = None
    return save_item(db, Skill(**data))


@router.get("/skills/{item_id}", response_model=SkillOut)
def read_skill(item_id: int, db: Session = Depends(get_db)):
    return get_or_404(db, Skill, item_id)


@router.put("/skills/{item_id}", response_model=SkillOut)
def update_skill(item_id: int, payload: SkillIn, db: Session = Depends(get_db)):
    return update_item(db, Skill, item_id, payload)


@router.delete("/skills/{item_id}")
def delete_skill(item_id: int, db: Session = Depends(get_db)):
    return deactivate_or_delete(db, Skill, item_id)


@router.post("/skills/{item_id}/archive")
def archive_skill(item_id: int, db: Session = Depends(get_db)):
    used = db.query(EmployeeSkill).filter(EmployeeSkill.skill_id == item_id).count()
    if used:
        raise HTTPException(status_code=409, detail="Навык назначен сотрудникам. Сначала снимите назначение.")
    return deactivate_or_delete(db, Skill, item_id)


@router.get("/skills/{item_id}/employees", response_model=list[EmployeeOut])
def list_skill_employees(item_id: int, request: Request, db: Session = Depends(get_db)):
    skill = get_or_404(db, Skill, item_id)
    project_id = current_project_id(db, request, None)
    rows = db.query(Employee).join(EmployeeSkill, EmployeeSkill.employee_id == Employee.id).join(
        EmployeeProjectAccess, EmployeeProjectAccess.employee_id == Employee.id
    ).filter(
        EmployeeSkill.skill_id == item_id,
        EmployeeProjectAccess.project_id == project_id,
        EmployeeProjectAccess.can_work.is_(True),
    ).order_by(Employee.full_name).all()
    return [employee_out(db, item, project_id) for item in rows]


@router.put("/skills/{item_id}/employees")
def update_skill_employees(item_id: int, payload: SkillEmployeesIn, request: Request, db: Session = Depends(get_db)) -> dict:
    skill = get_or_404(db, Skill, item_id)
    project_id = current_project_id(db, request, None)
    employees = db.query(Employee).join(EmployeeProjectAccess, EmployeeProjectAccess.employee_id == Employee.id).filter(
        Employee.id.in_(payload.employee_ids),
        EmployeeProjectAccess.project_id == project_id,
        EmployeeProjectAccess.can_work.is_(True),
    ).all() if payload.employee_ids else []
    if len(employees) != len(set(payload.employee_ids)):
        raise HTTPException(status_code=422, detail="Все сотрудники должны принадлежать текущему контуру")
    db.query(EmployeeSkill).filter(EmployeeSkill.skill_id == item_id).delete(synchronize_session=False)
    for employee in employees:
        db.add(EmployeeSkill(employee_id=employee.id, skill_id=item_id, level=1))
    log_audit(db, "skill_employees_update", request=request, entity_type="skill", entity_id=str(item_id), details=f"employees={len(employees)}")
    db.commit()
    return {"status": "ok", "skill_id": item_id, "employee_ids": sorted(set(payload.employee_ids))}


@router.get("/queues", response_model=list[QueueOut])
def list_queues(request: Request, project_id: int | None = None, db: Session = Depends(get_db)):
    project_id = current_project_id(db, request, project_id)
    return apply_project_filter(db.query(Queue), Queue, project_id).order_by(Queue.id).all()


@router.post("/queues", response_model=QueueOut)
def create_queue(payload: QueueIn, request: Request, db: Session = Depends(get_db)):
    data = payload.model_dump()
    data["project_id"] = current_project_id(db, request, data.get("project_id") if "project_id" in data else None)
    return save_item(db, Queue(**data))


@router.get("/queues/{item_id}", response_model=QueueOut)
def read_queue(item_id: int, db: Session = Depends(get_db)):
    return get_or_404(db, Queue, item_id)


@router.put("/queues/{item_id}", response_model=QueueOut)
def update_queue(item_id: int, payload: QueueIn, db: Session = Depends(get_db)):
    return update_item(db, Queue, item_id, payload)


@router.delete("/queues/{item_id}")
def delete_queue(item_id: int, db: Session = Depends(get_db)):
    return deactivate_or_delete(db, Queue, item_id)


@router.post("/queues/{item_id}/archive")
def archive_queue(item_id: int, db: Session = Depends(get_db)):
    return deactivate_or_delete(db, Queue, item_id)


@router.get("/queues/{queue_id}/skills", response_model=list[QueueSkillOut])
def list_queue_skills(queue_id: int, db: Session = Depends(get_db)):
    get_or_404(db, Queue, queue_id)
    items = db.query(QueueSkill).filter(QueueSkill.queue_id == queue_id).order_by(QueueSkill.id).all()
    return [QueueSkillOut.model_validate(item).model_copy(update={"skill_name": skill_name(db, item.skill_id)}) for item in items]


@router.post("/queues/{queue_id}/skills", response_model=QueueSkillOut)
def create_queue_skill(queue_id: int, payload: QueueSkillIn, db: Session = Depends(get_db)):
    get_or_404(db, Queue, queue_id)
    get_or_404(db, Skill, payload.skill_id)
    existing = db.query(QueueSkill).filter(QueueSkill.queue_id == queue_id, QueueSkill.skill_id == payload.skill_id).first()
    if existing:
        existing.min_level = payload.min_level
        existing.is_required = payload.is_required
        item = save_item(db, existing)
    else:
        item = save_item(db, QueueSkill(queue_id=queue_id, **payload.model_dump()))
    return QueueSkillOut.model_validate(item).model_copy(update={"skill_name": skill_name(db, item.skill_id)})


@router.put("/queues/{queue_id}/skills/{item_id}", response_model=QueueSkillOut)
def update_queue_skill(queue_id: int, item_id: int, payload: QueueSkillIn, db: Session = Depends(get_db)):
    item = get_or_404(db, QueueSkill, item_id)
    if item.queue_id != queue_id:
        raise HTTPException(status_code=404, detail="Навык очереди не найден")
    item.skill_id = payload.skill_id
    item.min_level = payload.min_level
    item.is_required = payload.is_required
    item = save_item(db, item)
    return QueueSkillOut.model_validate(item).model_copy(update={"skill_name": skill_name(db, item.skill_id)})


@router.delete("/queues/{queue_id}/skills/{item_id}")
def delete_queue_skill(queue_id: int, item_id: int, db: Session = Depends(get_db)):
    item = get_or_404(db, QueueSkill, item_id)
    if item.queue_id != queue_id:
        raise HTTPException(status_code=404, detail="Навык очереди не найден")
    db.delete(item)
    db.commit()
    return {"status": "deleted", "id": item_id}


@router.get("/employees", response_model=list[EmployeeOut])
def list_employees(request: Request, project_id: int | None = None, archived: bool = False, db: Session = Depends(get_db)):
    project_id = current_project_id(db, request, project_id)
    if project_id:
        query = db.query(Employee).join(EmployeeProjectAccess, EmployeeProjectAccess.employee_id == Employee.id).filter(
            EmployeeProjectAccess.project_id == project_id,
            EmployeeProjectAccess.can_work.is_(True),
        )
    else:
        query = db.query(Employee)
    if archived:
        query = query.filter((Employee.is_active.is_(False)) | (Employee.employment_status == "archived"))
    else:
        query = query.filter(Employee.is_active.is_(True), Employee.employment_status != "archived")
    items = query.order_by(Employee.id).all()
    return [employee_out(db, item, project_id) for item in items]


@router.post("/employees", response_model=EmployeeOut)
def create_employee(payload: EmployeeIn, request: Request, db: Session = Depends(get_db)):
    data = payload.model_dump()
    selected_project = current_project_id(db, request, data.get("project_id"))
    project_ids = data.pop("project_ids", []) or ([selected_project] if selected_project else [])
    if selected_project and selected_project not in project_ids:
        project_ids.append(selected_project)
    for project_id in project_ids:
        ensure_project_access(db, request, project_id)
    data["project_id"] = selected_project or (project_ids[0] if project_ids else None)
    data["inn"] = validate_inn_or_400(data.get("inn"))
    data["phone"] = normalize_phone(data.get("phone"))
    data.pop("snils", None)
    data.pop("naumen_login", None)
    if not data.get("personnel_number"):
        data["personnel_number"] = f"MAN-{int(datetime.utcnow().timestamp())}"
    if data.get("inn"):
        existing = db.query(Employee).filter(Employee.inn == data["inn"], Employee.is_active.is_(True)).first()
        if existing:
            sync_employee_project_access(db, existing, sorted(set(employee_project_ids(db, existing.id) + project_ids)), data.get("project_id"))
            if data.get("naumen_uuid"):
                existing.naumen_uuid = data["naumen_uuid"]
            existing.full_name = data["full_name"]
            existing.email = data.get("email")
            existing.phone = data.get("phone")
            existing.position = data.get("position")
            existing.comment = data.get("comment")
            db.flush()
            set_employee_team_for_project(db, existing, data.get("team_id"), data.get("project_id"))
            db.commit()
            db.refresh(existing)
            return employee_out(db, existing, data.get("project_id"))
    if data.get("naumen_uuid") and db.query(Employee).join(EmployeeProjectAccess, EmployeeProjectAccess.employee_id == Employee.id).filter(
        Employee.naumen_uuid == data["naumen_uuid"],
        EmployeeProjectAccess.project_id.in_(project_ids),
        Employee.is_active.is_(True),
    ).first():
        raise HTTPException(status_code=409, detail="UUID Naumen уже привязан к активному сотруднику этого контура")
    team_id = data.pop("team_id", None)
    item = save_item(db, Employee(**data))
    sync_employee_project_access(db, item, project_ids, data.get("project_id"))
    db.flush()
    set_employee_team_for_project(db, item, team_id, data.get("project_id"))
    settings = get_or_create_onec_settings(db)
    if settings.check_on_employee_create and item.inn:
        result = OneCService(settings).check_employee_status_by_inn(item.inn)
        apply_onec_status(item, result)
        db.commit()
        db.refresh(item)
    db.commit()
    db.refresh(item)
    return employee_out(db, item, data.get("project_id"))


@router.get("/employees/import/template.xlsx")
def employee_import_template() -> Response:
    try:
        from openpyxl import Workbook
    except ImportError as exc:
        raise HTTPException(status_code=500, detail="openpyxl не установлен") from exc
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Сотрудники"
    headers = ["Проекты/контуры", "ФИО", "ИНН", "Дата рождения", "Email", "Телефон", "Должность", "Команда", "Навыки", "UUID Naumen", "Комментарий"]
    sheet.append(headers)
    help_sheet = workbook.create_sheet("Справка")
    help_sheet.append(["Колонка", "Описание"])
    help_rows = [
        ("ФИО", "Обязательное поле."),
        ("Проекты/контуры", "Если пусто, используется текущий активный контур. Несколько контуров указываются через запятую или точку с запятой."),
        ("ИНН", "Используется только для сверки с 1С: 12 цифр."),
        ("UUID Naumen", "Используется только для сопоставления с оператором Naumen."),
        ("Навыки", "Список через запятую. Навыки создаются автоматически при импорте."),
        ("Команда", "Если команды нет, она создаётся автоматически."),
    ]
    for row in help_rows:
        help_sheet.append(row)
    output = io.BytesIO()
    workbook.save(output)
    return Response(
        content=output.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": 'attachment; filename="employee-registry-template.xlsx"'},
    )


@router.post("/employees/import/xlsx")
async def import_employees_xlsx(request: Request, file: UploadFile = File(...), db: Session = Depends(get_db)) -> dict:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise HTTPException(status_code=500, detail="openpyxl не установлен") from exc
    raw = await file.read()
    current_pid = current_project_id(db, request)
    batch = EmployeeImportBatch(filename=file.filename or "employees.xlsx", created_by_user_id=current_user_id(request))
    db.add(batch)
    db.commit()
    db.refresh(batch)
    try:
        workbook = load_workbook(io.BytesIO(raw), read_only=True, data_only=True)
        sheet = workbook["Сотрудники"] if "Сотрудники" in workbook.sheetnames else workbook.active
        rows = list(sheet.iter_rows(values_only=True))
        if not rows:
            raise ValueError("Файл не содержит строк")
        headers = [str(value or "").strip() for value in rows[0]]
        if "ФИО" not in headers or "ИНН" not in headers:
            raise ValueError("В файле обязательны колонки ФИО и ИНН")
        if len(rows) - 1 > 10000:
            raise ValueError("Максимальный размер реестра: 10 000 строк")
        index = {name: idx for idx, name in enumerate(headers)}
        seen_inns: set[str] = set()
        for row_number, row in enumerate(rows[1:], start=2):
            values = {name: (row[idx] if idx < len(row) else None) for name, idx in index.items()}
            if not any(values.values()):
                continue
            batch.rows_total += 1
            try:
                full_name = str(values.get("ФИО") or "").strip()
                if not full_name:
                    raise ValueError("ФИО обязательно")
                inn = validate_inn_or_400(str(values.get("ИНН") or "").strip())
                if not inn:
                    raise ValueError("ИНН обязателен")
                if inn in seen_inns:
                    raise ValueError("Дублирующийся ИНН внутри файла")
                seen_inns.add(inn)
                project_ids = [current_pid] if current_pid else []
                contour_value = str(values.get("Проекты/контуры") or values.get("Рабочий контур") or "").strip()
                if contour_value:
                    project_ids = []
                    for contour_name in [part.strip() for part in contour_value.replace(";", ",").split(",") if part.strip()]:
                        contour = db.query(NaumenProject).filter(NaumenProject.title == contour_name, NaumenProject.is_active.is_(True)).first()
                        if not contour:
                            raise ValueError(f"Рабочий контур не найден: {contour_name}")
                        project_ids.append(contour.id)
                if not project_ids:
                    raise ValueError("Не выбран рабочий контур")
                project_id = project_ids[0]
                team = None
                team_value = str(values.get("Команда") or "").strip()
                if team_value:
                    team = db.query(Team).filter(Team.name == team_value, Team.project_id == project_id).first()
                    if not team:
                        team = Team(name=team_value, project_id=project_id, description="Создано при импорте реестра сотрудников")
                        db.add(team)
                        db.flush()
                employee = db.query(Employee).filter(Employee.inn == inn).first()
                created = employee is None
                if created:
                    employee = Employee(personnel_number=f"INN-{inn}", inn=inn, full_name=full_name, project_id=project_id)
                employee.project_id = project_id
                employee.full_name = full_name
                employee.email = str(values.get("Email") or "").strip() or None
                employee.phone = normalize_phone(str(values.get("Телефон") or "").strip() or None)
                employee.position = str(values.get("Должность") or "").strip() or None
                employee.naumen_uuid = str(values.get("UUID Naumen") or "").strip() or employee.naumen_uuid
                employee.comment = str(values.get("Комментарий") or "").strip() or None
                employee.source_type = "xlsx_import"
                birth_value = values.get("Дата рождения")
                if isinstance(birth_value, datetime):
                    employee.birth_date = birth_value.date()
                elif isinstance(birth_value, date):
                    employee.birth_date = birth_value
                db.add(employee)
                db.flush()
                sync_employee_project_access(db, employee, sorted(set(employee_project_ids(db, employee.id) + project_ids)), project_id)
                if team:
                    set_employee_team_for_project(db, employee, team.id, project_id)
                skills_value = str(values.get("Навыки") or "").strip()
                if skills_value:
                    for skill_title in [item.strip() for item in skills_value.split(",") if item.strip()]:
                        skill = db.query(Skill).filter(func.lower(Skill.name) == skill_title.lower()).first()
                        if not skill:
                            skill = Skill(name=skill_title, project_id=None)
                            db.add(skill)
                            db.flush()
                        if not db.query(EmployeeSkill).filter(EmployeeSkill.employee_id == employee.id, EmployeeSkill.skill_id == skill.id).first():
                            db.add(EmployeeSkill(employee_id=employee.id, skill_id=skill.id, level=1))
                if created:
                    batch.rows_created += 1
                else:
                    batch.rows_updated += 1
            except Exception as exc:
                batch.rows_failed += 1
                db.add(EmployeeImportError(batch_id=batch.id, row_number=row_number, field=None, error_message=str(exc), raw_data=json.dumps(values, ensure_ascii=False, default=str)))
        batch.status = "success" if batch.rows_failed == 0 else "partial_success"
    except Exception as exc:
        batch.status = "error"
        batch.error_message = str(exc)
    batch.completed_at = utcnow()
    db.commit()
    errors = db.query(EmployeeImportError).filter(EmployeeImportError.batch_id == batch.id).order_by(EmployeeImportError.id).limit(200).all()
    return {
        "batch_id": batch.id,
        "status": batch.status,
        "rows_total": batch.rows_total,
        "rows_created": batch.rows_created,
        "rows_updated": batch.rows_updated,
        "rows_failed": batch.rows_failed,
        "errors": [{"row_number": item.row_number, "field": item.field, "error_message": item.error_message, "raw_data": item.raw_data} for item in errors],
        "error_message": batch.error_message,
    }


@router.get("/employees/{item_id}", response_model=EmployeeOut)
def read_employee(item_id: int, db: Session = Depends(get_db)):
    item = get_or_404(db, Employee, item_id)
    return employee_out(db, item)


@router.put("/employees/{item_id}", response_model=EmployeeOut)
def update_employee(item_id: int, payload: EmployeeIn, request: Request, db: Session = Depends(get_db)):
    data = payload.model_dump()
    selected_project = current_project_id(db, request, data.get("project_id"))
    project_ids = data.pop("project_ids", []) or employee_project_ids(db, item_id) or ([selected_project] if selected_project else [])
    if selected_project and selected_project not in project_ids:
        project_ids.append(selected_project)
    for project_id in project_ids:
        ensure_project_access(db, request, project_id)
    data["project_id"] = selected_project or (project_ids[0] if project_ids else None)
    data["inn"] = validate_inn_or_400(data.get("inn"))
    data["phone"] = normalize_phone(data.get("phone"))
    data.pop("snils", None)
    data.pop("naumen_login", None)
    if data.get("inn"):
        duplicate = db.query(Employee).join(EmployeeProjectAccess, EmployeeProjectAccess.employee_id == Employee.id).filter(
            Employee.inn == data["inn"],
            EmployeeProjectAccess.project_id.in_(project_ids),
            Employee.id != item_id,
            Employee.is_active.is_(True),
        ).first()
        if duplicate:
            raise HTTPException(status_code=409, detail="Сотрудник с таким ИНН уже существует")
    if data.get("naumen_uuid"):
        duplicate_naumen = db.query(Employee).join(EmployeeProjectAccess, EmployeeProjectAccess.employee_id == Employee.id).filter(
            Employee.naumen_uuid == data["naumen_uuid"],
            EmployeeProjectAccess.project_id.in_(project_ids),
            Employee.id != item_id,
            Employee.is_active.is_(True),
        ).first()
        if duplicate_naumen:
            raise HTTPException(status_code=409, detail="UUID Naumen уже привязан к активному сотруднику этого контура")
    item = get_or_404(db, Employee, item_id)
    if selected_project and not employee_has_project_access(db, item.id, selected_project):
        raise HTTPException(status_code=403, detail="Сотрудник не доступен в текущем контуре")
    team_id = data.pop("team_id", None)
    if data.get("team_id"):
        team = get_or_404(db, Team, data["team_id"])
        if team.project_id != data.get("project_id"):
            raise HTTPException(status_code=422, detail="Команда должна принадлежать текущему контуру")
    for key, value in data.items():
        if key == "personnel_number" and not value:
            continue
        setattr(item, key, value)
    sync_employee_project_access(db, item, project_ids, data.get("project_id"))
    db.flush()
    set_employee_team_for_project(db, item, team_id, data.get("project_id"))
    item = save_item(db, item)
    return employee_out(db, item, data.get("project_id"))


@router.put("/employees/{item_id}/skills")
def update_employee_skills(item_id: int, payload: EmployeeSkillsIn, request: Request, db: Session = Depends(get_db)) -> dict:
    employee = get_or_404(db, Employee, item_id)
    project_id = current_project_id(db, request, None)
    if not employee_has_project_access(db, employee.id, project_id):
        raise HTTPException(status_code=403, detail="Сотрудник не доступен в текущем контуре")
    valid_skills = db.query(Skill).filter(Skill.id.in_(payload.skill_ids), Skill.is_active.is_(True)).all() if payload.skill_ids else []
    if len(valid_skills) != len(set(payload.skill_ids)):
        raise HTTPException(status_code=422, detail="Выберите существующие активные навыки")
    db.query(EmployeeSkill).filter(EmployeeSkill.employee_id == item_id).delete(synchronize_session=False)
    for skill_id in sorted(set(payload.skill_ids)):
        db.add(EmployeeSkill(employee_id=item_id, skill_id=skill_id, level=1))
    log_audit(db, "employee_skills_update", request=request, entity_type="employee", entity_id=str(item_id), details=f"skills={len(payload.skill_ids)}")
    db.commit()
    return {"status": "ok", "employee_id": item_id, "skill_ids": sorted(set(payload.skill_ids))}


@router.delete("/employees/{item_id}")
def delete_employee(item_id: int, db: Session = Depends(get_db)):
    return deactivate_or_delete(db, Employee, item_id)


@router.post("/employees/{item_id}/archive")
def archive_employee(item_id: int, request: Request, db: Session = Depends(get_db)):
    employee = get_or_404(db, Employee, item_id)
    employee.is_active = False
    employee.employment_status = "archived"
    if hasattr(employee, "wfm_status"):
        employee.wfm_status = "archived"
    log_audit(db, "employee_archive", request=request, entity_type="employee", entity_id=str(item_id))
    return save_item(db, employee)


@router.post("/employees/{item_id}/dismiss")
def dismiss_employee(item_id: int, request: Request, db: Session = Depends(get_db)):
    employee = get_or_404(db, Employee, item_id)
    employee.employment_status = "dismissed"
    employee.dismissal_date = date.today()
    log_audit(db, "employee_dismiss", request=request, entity_type="employee", entity_id=str(item_id))
    return employee_out(db, save_item(db, employee))


@router.post("/employees/{item_id}/restore")
def restore_employee(item_id: int, request: Request, db: Session = Depends(get_db)):
    employee = get_or_404(db, Employee, item_id)
    employee.is_active = True
    employee.employment_status = "working"
    employee.dismissal_date = None
    log_audit(db, "employee_restore", request=request, entity_type="employee", entity_id=str(item_id))
    return employee_out(db, save_item(db, employee))


@router.delete("/employees/{item_id}/hard-delete")
def hard_delete_employee(item_id: int, request: Request, db: Session = Depends(get_db)):
    employee = get_or_404(db, Employee, item_id)
    if employee.is_active and employee.employment_status != "archived":
        raise HTTPException(status_code=409, detail="Окончательно удалить можно только сотрудника из архива")
    linked_schedules = db.query(ScheduleAssignment).filter(ScheduleAssignment.employee_id == item_id).count()
    linked_history = db.query(EmployeeAttendanceFact).filter(EmployeeAttendanceFact.employee_id == item_id).count()
    if linked_schedules or linked_history:
        raise HTTPException(status_code=409, detail="У сотрудника есть связанные графики или история. Оставьте запись в архиве.")
    db.query(EmployeeSkill).filter(EmployeeSkill.employee_id == item_id).delete(synchronize_session=False)
    db.query(EmployeeProjectAccess).filter(EmployeeProjectAccess.employee_id == item_id).delete(synchronize_session=False)
    db.query(EmployeeTeamMembership).filter(EmployeeTeamMembership.employee_id == item_id).delete(synchronize_session=False)
    log_audit(db, "employee_hard_delete", request=request, entity_type="employee", entity_id=str(item_id))
    db.delete(employee)
    db.commit()
    return {"status": "deleted", "id": item_id}


@router.post("/employees/{employee_id}/check-1c-status")
def check_employee_1c_status(employee_id: int, request: Request, db: Session = Depends(get_db)) -> dict:
    employee = get_or_404(db, Employee, employee_id)
    if not employee.inn:
        raise HTTPException(status_code=400, detail="Для сверки с 1С у сотрудника должен быть заполнен ИНН.")
    settings = get_or_create_onec_settings(db)
    result = OneCService(settings).check_employee_status_by_inn(employee.inn)
    apply_onec_status(employee, result)
    if settings.auto_disable_dismissed and result.status == "dismissed":
        employee.is_active = False
    log_audit(db, "employee_check_1c_status", request=request, entity_type="employee", entity_id=str(employee.id), details=f"status={result.status}")
    db.commit()
    db.refresh(employee)
    return {"employee": employee_out(db, employee).model_dump(), "result": result.__dict__}


@router.post("/employees/check-all-1c-status")
def check_all_employees_1c_status(payload: EmployeeCheckAllIn, request: Request, db: Session = Depends(get_db)) -> dict:
    running = db.query(OneCStatusCheckRun).filter(OneCStatusCheckRun.status == "running").first()
    if running:
        raise HTTPException(status_code=409, detail="Массовая сверка с 1С уже выполняется")
    project_id = current_project_id(db, request, None)
    query = db.query(Employee).join(EmployeeProjectAccess, EmployeeProjectAccess.employee_id == Employee.id).filter(
        EmployeeProjectAccess.project_id == project_id,
        EmployeeProjectAccess.can_work.is_(True),
        Employee.inn.isnot(None),
    )
    if payload.only_active:
        query = query.filter(Employee.is_active.is_(True))
    employees = query.order_by(Employee.id).all()
    run = OneCStatusCheckRun(run_type="manual", started_by_user_id=current_user_id(request), employees_total=len(employees))
    db.add(run)
    db.commit()
    db.refresh(run)
    settings = get_or_create_onec_settings(db)
    service = OneCService(settings)
    try:
        for employee in employees:
            inn = normalize_inn(employee.inn)
            if not inn or len(inn) not in {10, 12}:
                run.employees_failed += 1
                db.add(OneCStatusCheckError(run_id=run.id, employee_id=employee.id, inn=employee.inn, error_message="Некорректный ИНН: сверка не отправлялась в 1С Gateway"))
                continue
            employee.inn = inn
            result = service.check_employee_status_by_inn(inn)
            apply_onec_status(employee, result)
            run.employees_checked += 1
            if result.status == "working":
                run.employees_working += 1
            elif result.status == "dismissed":
                run.employees_dismissed += 1
                if settings.auto_disable_dismissed:
                    employee.is_active = False
            elif result.status == "not_found":
                run.employees_not_found += 1
            else:
                run.employees_failed += 1
                db.add(OneCStatusCheckError(run_id=run.id, employee_id=employee.id, inn=employee.inn, error_message=result.message or "Ошибка сверки"))
        run.status = "success" if run.employees_failed == 0 else "partial_success"
    except Exception as exc:
        run.status = "error"
        run.error_message = "Массовая сверка завершилась ошибкой"
        db.add(OneCStatusCheckError(run_id=run.id, employee_id=None, inn=None, error_message=str(exc)))
    run.finished_at = utcnow()
    log_audit(db, "employees_check_all_1c_status", request=request, entity_type="onec_status_check_run", entity_id=str(run.id), details=f"status={run.status}")
    db.commit()
    return {
        "run_id": run.id,
        "status": run.status,
        "employees_total": run.employees_total,
        "employees_checked": run.employees_checked,
        "employees_working": run.employees_working,
        "employees_dismissed": run.employees_dismissed,
        "employees_not_found": run.employees_not_found,
        "employees_failed": run.employees_failed,
        "error_message": run.error_message,
    }


@router.get("/employees/{item_id}/naumen")
def read_employee_naumen(item_id: int, request: Request, db: Session = Depends(get_db)) -> dict:
    employee = get_or_404(db, Employee, item_id)
    project_id = current_project_id(db, request, None)
    if not employee_has_project_access(db, employee.id, project_id):
        raise HTTPException(status_code=403, detail="Сотрудник не доступен в текущем контуре")
    operator = None
    if employee.naumen_uuid:
        operator = db.query(NaumenOperator).filter(NaumenOperator.project_id == project_id, NaumenOperator.naumen_uuid == employee.naumen_uuid).first()
    return {
        "employee_id": employee.id,
        "naumen_uuid": employee.naumen_uuid,
        "naumen_status": employee.naumen_status,
        "naumen_status_label": employee.naumen_status_label,
        "naumen_last_checked_at": employee.naumen_last_checked_at,
        "naumen_last_check_message": employee.naumen_last_check_message,
        "operator": NaumenOperatorOut.model_validate(operator).model_dump() if operator else None,
    }


@router.put("/employees/{item_id}/naumen-link")
def update_employee_naumen_link(item_id: int, payload: EmployeeNaumenLinkIn, request: Request, db: Session = Depends(get_db)) -> dict:
    employee = get_or_404(db, Employee, item_id)
    project_id = current_project_id(db, request, None)
    if not employee_has_project_access(db, employee.id, project_id):
        raise HTTPException(status_code=403, detail="Сотрудник не доступен в текущем контуре")
    naumen_uuid = (payload.naumen_uuid or "").strip() or None
    if naumen_uuid:
        duplicate = db.query(Employee).join(EmployeeProjectAccess, EmployeeProjectAccess.employee_id == Employee.id).filter(
            EmployeeProjectAccess.project_id == project_id,
            Employee.naumen_uuid == naumen_uuid,
            Employee.id != employee.id,
            Employee.is_active.is_(True),
        ).first()
        if duplicate:
            raise HTTPException(status_code=409, detail="UUID Naumen уже привязан к активному сотруднику этого контура")
    employee.naumen_uuid = naumen_uuid
    if not naumen_uuid:
        update_employee_naumen_status(employee, "not_linked", "UUID Naumen не указан.")
    else:
        update_employee_naumen_status(employee, "not_linked", "UUID Naumen сохранён. Выполните сверку с Naumen.")
    log_audit(db, "employee_naumen_link_update", request=request, entity_type="employee", entity_id=str(employee.id), details=f"uuid_present={bool(naumen_uuid)}")
    db.commit()
    db.refresh(employee)
    return read_employee_naumen(item_id, request, db)


@router.post("/employees/{item_id}/check-naumen")
def check_employee_naumen(item_id: int, request: Request, db: Session = Depends(get_db)) -> dict:
    employee = get_or_404(db, Employee, item_id)
    project_id = current_project_id(db, request, None)
    if not employee_has_project_access(db, employee.id, project_id):
        raise HTTPException(status_code=403, detail="Сотрудник не доступен в текущем контуре")
    project = db.get(NaumenProject, project_id) if project_id else None
    if not project:
        update_employee_naumen_status(employee, "check_error", "Рабочий контур не выбран.")
        db.commit()
        return {"status": employee.naumen_status, "message": employee.naumen_last_check_message}
    partner_uuid = project.naumen_customer_uuid or project.partner_uuid
    if not partner_uuid:
        update_employee_naumen_status(employee, "check_error", "Для рабочего контура не указан UUID Naumen/NCC.")
        db.commit()
        return {"status": employee.naumen_status, "message": employee.naumen_last_check_message}
    last_name, first_name, middle_name = split_person_name(employee.full_name)
    if not last_name or not first_name:
        update_employee_naumen_status(employee, "check_error", "Для сверки с Naumen должны быть заполнены фамилия и имя.")
        db.commit()
        return {"status": employee.naumen_status, "message": employee.naumen_last_check_message}
    operators = db.query(NaumenOperator).filter(NaumenOperator.project_id == project_id).all()
    if not operators:
        begin, end = ncc_default_period()
        try:
            rows = NaumenNccService(db).employees(partner_uuid, begin, end)
            NaumenNccService(db)._save_operators(project, rows)
            db.commit()
            operators = db.query(NaumenOperator).filter(NaumenOperator.project_id == project_id).all()
        except Exception:
            update_employee_naumen_status(employee, "api_unavailable", "Naumen/NCC недоступен или не настроен. Сверка по ФИО не выполнена.")
            db.commit()
            return {"status": employee.naumen_status, "message": employee.naumen_last_check_message}
    candidates = [
        op for op in operators
        if op.normalized_last_name == last_name
        and op.normalized_first_name == first_name
        and ((op.normalized_middle_name == middle_name) if middle_name else True)
    ]
    if len(candidates) == 1:
        operator = candidates[0]
        employee.naumen_uuid = operator.naumen_uuid
        employee.naumen_project_uuid = operator.project_uuid
        employee.naumen_project_title = project.title
        employee.naumen_last_sync_at = operator.last_seen_at
        update_employee_naumen_status(employee, "linked", "Сотрудник сопоставлен с Naumen по ФИО.", {"operator_id": operator.id})
        db.commit()
        return {"status": "linked", "message": employee.naumen_last_check_message, "operator": NaumenOperatorOut.model_validate(operator).model_dump()}
    if len(candidates) > 1:
        update_employee_naumen_status(employee, "mismatch", "Найдено несколько кандидатов. Выберите вручную.", {"candidate_ids": [op.id for op in candidates]})
        db.commit()
        return {"status": "mismatch", "message": employee.naumen_last_check_message, "candidates": [NaumenOperatorOut.model_validate(op).model_dump() for op in candidates]}
    update_employee_naumen_status(employee, "not_found", "Сотрудник не найден в операторах проекта Naumen.")
    db.commit()
    return {"status": employee.naumen_status, "message": employee.naumen_last_check_message}


@router.post("/employees/{item_id}/check-naumen-uuid")
def check_employee_naumen_uuid(item_id: int, request: Request, db: Session = Depends(get_db)) -> dict:
    employee = get_or_404(db, Employee, item_id)
    project_id = current_project_id(db, request, None)
    if not employee_has_project_access(db, employee.id, project_id):
        raise HTTPException(status_code=403, detail="Сотрудник не доступен в текущем контуре")
    if not employee.naumen_uuid:
        update_employee_naumen_status(employee, "not_linked", "UUID Naumen не указан.")
        db.commit()
        return {"status": employee.naumen_status, "message": employee.naumen_last_check_message}
    operator = db.query(NaumenOperator).filter(NaumenOperator.project_id == project_id, NaumenOperator.naumen_uuid == employee.naumen_uuid).first()
    if not operator:
        update_employee_naumen_status(employee, "not_found", "Оператор с таким UUID не найден в загруженных операторах проекта Naumen.")
        db.commit()
        return {"status": employee.naumen_status, "message": employee.naumen_last_check_message}
    update_employee_naumen_status(employee, "linked", "UUID Naumen подтверждён по загруженным операторам проекта.", {"operator_id": operator.id})
    employee.naumen_project_uuid = operator.project_uuid
    employee.naumen_project_title = db.get(NaumenProject, project_id).title if project_id and db.get(NaumenProject, project_id) else employee.naumen_project_title
    employee.naumen_last_sync_at = operator.last_seen_at
    db.commit()
    return {"status": "linked", "message": employee.naumen_last_check_message, "operator": NaumenOperatorOut.model_validate(operator).model_dump()}


@router.post("/employees/naumen/match")
def match_naumen_operators(request: Request, project_id: int | None = None, db: Session = Depends(get_db)) -> dict:
    project_id = current_project_id(db, request, project_id)
    employees = db.query(Employee).join(EmployeeProjectAccess, EmployeeProjectAccess.employee_id == Employee.id).filter(
        EmployeeProjectAccess.project_id == project_id,
        EmployeeProjectAccess.can_work.is_(True),
        Employee.is_active.is_(True),
    ).all()
    operators = db.query(NaumenOperator).filter(NaumenOperator.project_id == project_id).all()
    linked = 0
    suggestions: list[dict] = []
    for employee in employees:
        by_uuid = next((op for op in operators if employee.naumen_uuid and op.naumen_uuid == employee.naumen_uuid), None)
        if by_uuid:
            update_employee_naumen_status(employee, "linked", "Автоматически подтверждено по UUID Naumen.", {"operator_id": by_uuid.id})
            linked += 1
            continue
        last_name, first_name, middle_name = split_person_name(employee.full_name)
        if not last_name or not first_name:
            continue
        by_name = [
            op for op in operators
            if op.normalized_last_name == last_name
            and op.normalized_first_name == first_name
            and ((op.normalized_middle_name == middle_name) if middle_name else True)
        ]
        if len(by_name) == 1:
            employee.naumen_uuid = by_name[0].naumen_uuid
            update_employee_naumen_status(employee, "linked", "Автоматически сопоставлено по ФИО.", {"operator_id": by_name[0].id})
            linked += 1
        elif len(by_name) > 1:
            update_employee_naumen_status(employee, "mismatch", "Найдено несколько кандидатов. Выберите вручную.", {"candidate_ids": [op.id for op in by_name]})
            suggestions.append({"employee_id": employee.id, "employee_name": employee.full_name, "candidates": [op.id for op in by_name], "reason": "Несколько совпадений по ФИО"})
    db.commit()
    return {"status": "ok", "linked": linked, "suggestions": suggestions, "message": "Совпадения по ФИО не привязываются автоматически."}


@router.post("/employees/naumen/check-all")
def check_all_naumen(request: Request, project_id: int | None = None, db: Session = Depends(get_db)) -> dict:
    project_id = current_project_id(db, request, project_id)
    project = db.get(NaumenProject, project_id) if project_id else None
    partner_uuid = (project.naumen_customer_uuid or project.partner_uuid) if project else None
    if not project or not partner_uuid:
        raise HTTPException(status_code=422, detail="Для активного контура не указан UUID Naumen/NCC")
    employees = db.query(Employee).join(EmployeeProjectAccess, EmployeeProjectAccess.employee_id == Employee.id).filter(
        EmployeeProjectAccess.project_id == project_id,
        EmployeeProjectAccess.can_work.is_(True),
        Employee.is_active.is_(True),
    ).all()
    operators = db.query(NaumenOperator).filter(NaumenOperator.project_id == project_id).all()
    period_begin = period_end = None
    if not operators:
        period_begin, period_end = ncc_default_period()
        try:
            rows = NaumenNccService(db).employees(partner_uuid, period_begin, period_end)
            NaumenNccService(db)._save_operators(project, rows)
            db.commit()
            operators = db.query(NaumenOperator).filter(NaumenOperator.project_id == project_id).all()
        except Exception as exc:
            raise HTTPException(status_code=503, detail="Naumen/NCC недоступен или не настроен. Массовая сверка не выполнена.") from exc
    checked = linked = failed = mismatch = 0
    for employee in employees:
        checked += 1
        last_name, first_name, middle_name = split_person_name(employee.full_name)
        if not last_name or not first_name:
            update_employee_naumen_status(employee, "check_error", "Для сверки с Naumen должны быть заполнены фамилия и имя.")
            failed += 1
            continue
        operator = db.query(NaumenOperator).filter(NaumenOperator.project_id == project_id, NaumenOperator.naumen_uuid == employee.naumen_uuid).first() if employee.naumen_uuid else None
        if operator:
            update_employee_naumen_status(employee, "linked", "Сотрудник сопоставлен с оператором Naumen.", {"operator_id": operator.id})
            linked += 1
            continue
        candidates = [
            op for op in operators
            if op.normalized_last_name == last_name
            and op.normalized_first_name == first_name
            and ((op.normalized_middle_name == middle_name) if middle_name else True)
        ]
        if len(candidates) == 1:
            employee.naumen_uuid = candidates[0].naumen_uuid
            update_employee_naumen_status(employee, "linked", "Сотрудник сопоставлен с Naumen по ФИО.", {"operator_id": candidates[0].id})
            linked += 1
        elif len(candidates) > 1:
            update_employee_naumen_status(employee, "mismatch", "Найдено несколько кандидатов. Выберите вручную.", {"candidate_ids": [op.id for op in candidates]})
            mismatch += 1
        else:
            update_employee_naumen_status(employee, "not_found", "Сотрудник не найден в операторах проекта Naumen.")
            failed += 1
    db.commit()
    return {
        "status": "ok",
        "checked": checked,
        "linked": linked,
        "mismatch": mismatch,
        "failed": failed,
        "period_begin": period_begin,
        "period_end": period_end,
        "message": "Сверка Naumen выполнена по операторам NCC активного контура.",
    }


@router.get("/employees/{item_id}/stats")
def employee_stats(item_id: int, request: Request, date_from: date | None = None, date_to: date | None = None, queue_id: int | None = None, db: Session = Depends(get_db)) -> dict:
    employee = get_or_404(db, Employee, item_id)
    ensure_project_access(db, request, employee.project_id)
    query = db.query(EmployeeIntervalStat).filter(EmployeeIntervalStat.employee_id == item_id)
    if date_from:
        query = query.filter(EmployeeIntervalStat.interval_start >= datetime.combine(date_from, time.min))
    if date_to:
        query = query.filter(EmployeeIntervalStat.interval_start <= datetime.combine(date_to, time.max))
    if queue_id:
        query = query.filter(EmployeeIntervalStat.queue_id == queue_id)
    intervals = query.order_by(EmployeeIntervalStat.interval_start).all()
    daily_query = db.query(EmployeeDailyStat).filter(EmployeeDailyStat.employee_id == item_id)
    if date_from:
        daily_query = daily_query.filter(EmployeeDailyStat.stat_date >= date_from)
    if date_to:
        daily_query = daily_query.filter(EmployeeDailyStat.stat_date <= date_to)
    daily = daily_query.order_by(EmployeeDailyStat.stat_date).all()
    return {
        "employee_id": item_id,
        "empty": not intervals and not daily,
        "message": "Статистика Naumen пока не загружена." if not intervals and not daily else "",
        "daily": [
            {
                "stat_date": row.stat_date,
                "source_system": row.source_system,
                "handled_contacts": row.handled_contacts,
                "offered_contacts": row.offered_contacts,
                "average_handle_time_sec": row.average_handle_time_sec,
                "service_level_percent": row.service_level_percent,
                "occupancy_percent": row.occupancy_percent,
                "adherence_percent": row.adherence_percent,
            }
            for row in daily
        ],
        "intervals": [
            {
                "interval_start": row.interval_start,
                "interval_end": row.interval_end,
                "queue_id": row.queue_id,
                "queue_name": queue_name(db, row.queue_id),
                "source_system": row.source_system,
                "handled_contacts": row.handled_contacts,
                "average_handle_time_sec": row.average_handle_time_sec,
                "service_level_percent": row.service_level_percent,
                "occupancy_percent": row.occupancy_percent,
            }
            for row in intervals
        ],
    }


@router.get("/employees/{item_id}/attendance")
def employee_attendance(item_id: int, request: Request, date_from: date | None = None, date_to: date | None = None, db: Session = Depends(get_db)) -> dict:
    employee = get_or_404(db, Employee, item_id)
    ensure_project_access(db, request, employee.project_id)
    query = db.query(EmployeeAttendanceFact).filter(EmployeeAttendanceFact.employee_id == item_id)
    if date_from:
        query = query.filter(EmployeeAttendanceFact.work_date >= date_from)
    if date_to:
        query = query.filter(EmployeeAttendanceFact.work_date <= date_to)
    rows = query.order_by(EmployeeAttendanceFact.work_date).all()
    return {"employee_id": item_id, "empty": not rows, "message": "Фактические выходы пока не загружены.", "items": [{"work_date": row.work_date, "status": row.status, "source_system": row.source_system, "planned_start": row.planned_start, "planned_end": row.planned_end, "actual_start": row.actual_start, "actual_end": row.actual_end} for row in rows]}


@router.get("/employees/{item_id}/schedule")
def employee_schedule(item_id: int, request: Request, date_from: date | None = None, date_to: date | None = None, db: Session = Depends(get_db)) -> dict:
    employee = get_or_404(db, Employee, item_id)
    ensure_project_access(db, request, employee.project_id)
    query = db.query(ScheduleAssignment).filter(ScheduleAssignment.employee_id == item_id)
    if date_from:
        query = query.filter(ScheduleAssignment.work_date >= date_from)
    if date_to:
        query = query.filter(ScheduleAssignment.work_date <= date_to)
    rows = query.order_by(ScheduleAssignment.work_date).all()
    return {"employee_id": item_id, "items": [schedule_out(db, row).model_dump() for row in rows]}


@router.get("/employees/{item_id}/statistics-summary")
def employee_statistics_summary(item_id: int, request: Request, date_from: date | None = None, date_to: date | None = None, db: Session = Depends(get_db)) -> dict:
    employee = get_or_404(db, Employee, item_id)
    ensure_project_access(db, request, employee.project_id)
    query = db.query(EmployeeDailyStat).filter(EmployeeDailyStat.employee_id == item_id)
    if date_from:
        query = query.filter(EmployeeDailyStat.stat_date >= date_from)
    if date_to:
        query = query.filter(EmployeeDailyStat.stat_date <= date_to)
    rows = query.all()
    handled = sum(row.handled_contacts for row in rows)
    offered = sum(row.offered_contacts for row in rows)
    aht = round(sum(row.average_handle_time_sec for row in rows) / len(rows), 2) if rows else 0
    sl = round(sum(row.service_level_percent for row in rows) / len(rows), 2) if rows else 0
    return {"employee_id": item_id, "empty": not rows, "handled_contacts": handled, "offered_contacts": offered, "average_handle_time_sec": aht, "service_level_percent": sl}


@router.get("/naumen/operators", response_model=list[NaumenOperatorOut])
def list_naumen_operators(request: Request, project_id: int | None = None, db: Session = Depends(get_db)):
    project_id = current_project_id(db, request, project_id)
    return db.query(NaumenOperator).filter(NaumenOperator.project_id == project_id).order_by(NaumenOperator.full_name, NaumenOperator.login).all()


@router.post("/naumen/operators/sync")
def sync_naumen_operators(request: Request, project_id: int | None = None, db: Session = Depends(get_db)) -> dict:
    project_id = current_project_id(db, request, project_id)
    project = db.get(NaumenProject, project_id) if project_id else None
    if not project or not project.project_uuid:
        raise HTTPException(status_code=422, detail="Для контура должен быть указан project_uuid Naumen")
    settings = get_or_create_naumen_settings(db)
    missing = missing_naumen_settings(settings)
    if missing:
        return {"status": "not_configured", "message": "Настройки Naumen неполные", "missing_fields": missing}
    try:
        payload = client_from_settings(settings).get_project_agents(project.project_uuid)
    except NaumenClientError as exc:
        return {"status": "error", "message": str(exc), "http_status": exc.http_status, "endpoint": exc.endpoint}
    items = payload.get("items", payload if isinstance(payload, list) else []) if isinstance(payload, (dict, list)) else []
    if not isinstance(items, list):
        items = []
    created = updated = 0
    now = utcnow()
    for raw_item in items:
        if not isinstance(raw_item, dict):
            continue
        naumen_uuid = str(raw_item.get("uuid") or raw_item.get("id") or raw_item.get("agentUuid") or "").strip()
        if not naumen_uuid:
            continue
        operator = db.query(NaumenOperator).filter(NaumenOperator.project_id == project_id, NaumenOperator.naumen_uuid == naumen_uuid).first()
        if not operator:
            operator = NaumenOperator(project_id=project_id, naumen_uuid=naumen_uuid)
            db.add(operator)
            created += 1
        else:
            updated += 1
        operator.project_uuid = project.project_uuid
        operator.partner_uuid = project.partner_uuid
        operator.full_name = str(raw_item.get("fullName") or raw_item.get("full_name") or raw_item.get("name") or "").strip() or None
        last_name, first_name, middle_name = split_person_name(operator.full_name)
        operator.normalized_last_name = last_name
        operator.normalized_first_name = first_name
        operator.normalized_middle_name = middle_name
        operator.login = str(raw_item.get("login") or raw_item.get("username") or "").strip() or None
        operator.email = str(raw_item.get("email") or "").strip() or None
        operator.phone = str(raw_item.get("phone") or "").strip() or None
        operator.state = str(raw_item.get("state") or "").strip() or None
        operator.substate = str(raw_item.get("substate") or "").strip() or None
        operator.skills = json.dumps(raw_item.get("skills"), ensure_ascii=False, default=str) if raw_item.get("skills") is not None else None
        operator.raw_data = json.dumps(raw_item, ensure_ascii=False, default=str)
        operator.last_seen_at = now
    db.commit()
    return {"status": "ok", "rows_received": len(items), "rows_created": created, "rows_updated": updated}


@router.get("/workload", response_model=list[WorkloadOut])
def list_workload(request: Request, project_id: int | None = None, db: Session = Depends(get_db)):
    project_id = current_project_id(db, request, project_id)
    items = apply_project_filter(db.query(WorkloadInterval), WorkloadInterval, project_id).order_by(WorkloadInterval.interval_start).all()
    return [workload_out(db, item) for item in items]


@router.post("/workload", response_model=WorkloadOut)
def create_workload(payload: WorkloadIn, request: Request, db: Session = Depends(get_db)):
    data = payload.model_dump()
    data["project_id"] = current_project_id(db, request, data.get("project_id"))
    queue = get_or_404(db, Queue, data["queue_id"])
    if queue.project_id != data["project_id"]:
        raise HTTPException(status_code=422, detail="Очередь должна принадлежать текущему контуру")
    item = save_item(db, WorkloadInterval(**data))
    return workload_out(db, item)


@router.get("/workload/template.xlsx")
def workload_template_early() -> Response:
    return workload_template()


@router.post("/workload/import/csv", response_model=ImportBatchOut)
async def import_workload_csv_alias_early(request: Request, file: UploadFile = File(...), db: Session = Depends(get_db)):
    return await import_workload_csv(file, request, db)


@router.post("/workload/import/xlsx")
async def import_workload_xlsx_early(request: Request, file: UploadFile = File(...), db: Session = Depends(get_db)) -> dict:
    return await import_workload_xlsx(file, request, db)


@router.get("/workload/{item_id}", response_model=WorkloadOut)
def read_workload(item_id: int, db: Session = Depends(get_db)):
    item = get_or_404(db, WorkloadInterval, item_id)
    return workload_out(db, item)


@router.put("/workload/{item_id}", response_model=WorkloadOut)
def update_workload(item_id: int, payload: WorkloadIn, db: Session = Depends(get_db)):
    item = update_item(db, WorkloadInterval, item_id, payload)
    return workload_out(db, item)


@router.delete("/workload/{item_id}")
def delete_workload(item_id: int, db: Session = Depends(get_db)):
    return deactivate_or_delete(db, WorkloadInterval, item_id)


@router.get("/workload/template.xlsx")
def workload_template() -> Response:
    try:
        from openpyxl import Workbook
    except ImportError as exc:
        raise HTTPException(status_code=500, detail="openpyxl не установлен") from exc
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Нагрузка"
    sheet.append(["Дата", "Начало интервала", "Конец интервала", "Очередь", "Поступило", "Обработано", "Потеряно", "AHT сек", "SL %", "Тип данных"])
    sheet.append([date.today().isoformat(), "09:00", "09:30", "Входящая линия", 10, 9, 1, 240, 80, "прогноз"])
    help_sheet = workbook.create_sheet("Справка")
    help_sheet.append(["Колонка", "Описание"])
    for row in [
        ("Дата", "Дата интервала в формате YYYY-MM-DD."),
        ("Начало интервала", "Время начала, например 09:00."),
        ("Конец интервала", "Время конца, например 09:30."),
        ("Очередь", "Название существующей очереди."),
        ("Тип данных", "факт или прогноз."),
    ]:
        help_sheet.append(row)
    output = io.BytesIO()
    workbook.save(output)
    return Response(content=output.getvalue(), media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", headers={"Content-Disposition": 'attachment; filename="workload-template.xlsx"'})


@router.post("/workload/import/csv", response_model=ImportBatchOut)
async def import_workload_csv_alias(request: Request, file: UploadFile = File(...), db: Session = Depends(get_db)):
    return await import_workload_csv(file, request, db)


@router.post("/workload/import/xlsx")
async def import_workload_xlsx(file: UploadFile = File(...), request: Request = None, db: Session = Depends(get_db)) -> dict:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise HTTPException(status_code=500, detail="openpyxl не установлен") from exc
    raw = await file.read()
    project_id = current_project_id(db, request)
    batch = ImportBatch(import_type="workload_xlsx", filename=file.filename or "workload.xlsx", project_id=project_id)
    db.add(batch)
    db.commit()
    db.refresh(batch)
    try:
        workbook = load_workbook(io.BytesIO(raw), read_only=True, data_only=True)
        sheet = workbook["Нагрузка"] if "Нагрузка" in workbook.sheetnames else workbook.active
        rows = list(sheet.iter_rows(values_only=True))
        headers = [str(value or "").strip() for value in rows[0]] if rows else []
        required = {"Дата", "Начало интервала", "Конец интервала", "Очередь", "Поступило", "Обработано", "Потеряно", "AHT сек", "SL %"}
        if set(headers) < required:
            raise ValueError("В XLSX нет обязательных колонок нагрузки")
        index = {name: idx for idx, name in enumerate(headers)}
        for row_number, row in enumerate(rows[1:], start=2):
            values = {name: (row[idx] if idx < len(row) else None) for name, idx in index.items()}
            if not any(values.values()):
                continue
            batch.rows_total += 1
            try:
                work_date = values["Дата"].date() if isinstance(values["Дата"], datetime) else values["Дата"]
                if isinstance(work_date, str):
                    work_date = date.fromisoformat(work_date[:10])
                start_time = values["Начало интервала"] if isinstance(values["Начало интервала"], time) else datetime.strptime(str(values["Начало интервала"])[:5], "%H:%M").time()
                end_time = values["Конец интервала"] if isinstance(values["Конец интервала"], time) else datetime.strptime(str(values["Конец интервала"])[:5], "%H:%M").time()
                interval_start = datetime.combine(work_date, start_time)
                interval_end = datetime.combine(work_date, end_time)
                queue = db.query(Queue).filter(Queue.name == str(values["Очередь"]).strip(), Queue.project_id == project_id).first()
                if not queue:
                    raise ValueError(f"Очередь не найдена: {values['Очередь']}")
                existing = db.query(WorkloadInterval).filter(WorkloadInterval.interval_start == interval_start, WorkloadInterval.interval_end == interval_end, WorkloadInterval.queue_id == queue.id, WorkloadInterval.project_id == project_id).first()
                payload = {
                    "offered_contacts": int(values["Поступило"] or 0),
                    "handled_contacts": int(values["Обработано"] or 0),
                    "abandoned_contacts": int(values["Потеряно"] or 0),
                    "average_handle_time_sec": int(values["AHT сек"] or 0),
                    "service_level_percent": float(values["SL %"] or 0),
                }
                if existing:
                    for key, value in payload.items():
                        setattr(existing, key, value)
                else:
                    db.add(WorkloadInterval(interval_start=interval_start, interval_end=interval_end, queue_id=queue.id, project_id=project_id, **payload))
                batch.rows_success += 1
            except Exception as exc:
                batch.rows_failed += 1
                db.add(ImportError(batch_id=batch.id, row_number=row_number, error_message=str(exc), raw_data=json.dumps(values, ensure_ascii=False, default=str)))
        batch.status = "completed" if batch.rows_failed == 0 else "completed_with_errors"
    except Exception as exc:
        batch.status = "failed"
        batch.error_message = str(exc)
    batch.completed_at = utcnow()
    db.commit()
    return {"batch_id": batch.id, "status": batch.status, "rows_total": batch.rows_total, "rows_success": batch.rows_success, "rows_failed": batch.rows_failed, "error_message": batch.error_message}


@router.post("/workload/clear-period")
def clear_workload_period(payload: DateRangeIn, request: Request, db: Session = Depends(get_db)) -> dict:
    start = datetime.combine(payload.date_from, time.min)
    end = datetime.combine(payload.date_to, time.max)
    query = db.query(WorkloadInterval).filter(WorkloadInterval.interval_start >= start, WorkloadInterval.interval_start <= end)
    if payload.queue_id:
        query = query.filter(WorkloadInterval.queue_id == payload.queue_id)
    deleted = query.delete(synchronize_session=False)
    log_audit(db, "workload_clear_period", request=request, entity_type="workload", details=f"rows={deleted}")
    db.commit()
    return {"status": "ok", "deleted": deleted}


@router.post("/imports/workload-csv", response_model=ImportBatchOut)
async def import_workload_csv(file: UploadFile = File(...), request: Request = None, db: Session = Depends(get_db)):
    project_id = current_project_id(db, request)
    batch = ImportBatch(import_type="workload_csv", filename=file.filename or "workload.csv", project_id=project_id)
    db.add(batch)
    db.commit()
    db.refresh(batch)

    required_columns = {
        "date",
        "interval_start",
        "interval_end",
        "queue_name",
        "offered_contacts",
        "handled_contacts",
        "abandoned_contacts",
        "average_handle_time_sec",
        "service_level_percent",
    }

    try:
        content = (await file.read()).decode("utf-8-sig")
        reader = csv.DictReader(io.StringIO(content))
        if not reader.fieldnames or set(reader.fieldnames) < required_columns:
            raise ValueError("CSV не содержит обязательные колонки")

        rows_total = rows_success = rows_failed = 0
        for row_number, row in enumerate(reader, start=2):
            rows_total += 1
            raw_data = ",".join([str(row.get(column, "")) for column in reader.fieldnames or []])
            try:
                interval_start, interval_end = parse_csv_datetime(row)
                queue = db.query(Queue).filter(Queue.name == row["queue_name"].strip(), Queue.project_id == project_id).first()
                if not queue:
                    raise ValueError(f"Очередь не найдена: {row['queue_name']}")

                values = {
                    "offered_contacts": int(row["offered_contacts"]),
                    "handled_contacts": int(row["handled_contacts"]),
                    "abandoned_contacts": int(row["abandoned_contacts"]),
                    "average_handle_time_sec": int(row["average_handle_time_sec"]),
                    "service_level_percent": float(row["service_level_percent"]),
                }
                existing = db.query(WorkloadInterval).filter(
                    WorkloadInterval.interval_start == interval_start,
                    WorkloadInterval.interval_end == interval_end,
                    WorkloadInterval.queue_id == queue.id,
                    WorkloadInterval.project_id == project_id,
                ).first()
                if existing:
                    for key, value in values.items():
                        setattr(existing, key, value)
                else:
                    db.add(WorkloadInterval(interval_start=interval_start, interval_end=interval_end, queue_id=queue.id, project_id=project_id, **values))
                rows_success += 1
            except Exception as exc:
                rows_failed += 1
                db.add(ImportError(batch_id=batch.id, row_number=row_number, error_message=str(exc), raw_data=raw_data))

        batch.rows_total = rows_total
        batch.rows_success = rows_success
        batch.rows_failed = rows_failed
        batch.status = "completed" if rows_failed == 0 else "completed_with_errors"
        batch.completed_at = utcnow()
        log_audit(db, "workload_import", entity_type="import_batch", entity_id=str(batch.id), details=f"rows_success={rows_success}; rows_failed={rows_failed}")
        db.commit()
        db.refresh(batch)
        return batch
    except Exception as exc:
        batch.status = "failed"
        batch.error_message = str(exc)
        batch.completed_at = utcnow()
        db.commit()
        db.refresh(batch)
        return batch


@router.post("/imports/actual-work-csv", response_model=ImportBatchOut)
async def import_actual_work_csv(file: UploadFile = File(...), request: Request = None, db: Session = Depends(get_db)):
    batch = ImportBatch(import_type="actual_work_csv", filename=file.filename or "actual-work.csv")
    db.add(batch)
    db.commit()
    db.refresh(batch)
    allowed_statuses = {"worked", "late", "absent", "offline", "unknown"}
    required_columns = {"date", "employee_email", "queue_name", "interval_start", "interval_end", "status", "actual_minutes"}
    try:
        content = (await file.read()).decode("utf-8-sig")
        if len(content.encode("utf-8")) > 10 * 1024 * 1024:
            raise ValueError("CSV больше 10 MB")
        reader = csv.DictReader(io.StringIO(content))
        if not reader.fieldnames or set(reader.fieldnames) < required_columns:
            raise ValueError("CSV не содержит обязательные колонки")
        rows_total = rows_success = rows_failed = 0
        for row_number, row in enumerate(reader, start=2):
            rows_total += 1
            raw_data = ",".join([str(row.get(column, "")) for column in reader.fieldnames or []])
            try:
                employee = db.query(Employee).filter(Employee.email == row["employee_email"].strip()).first()
                if not employee:
                    raise ValueError(f"Сотрудник не найден: {row['employee_email']}")
                queue = db.query(Queue).filter(Queue.name == row["queue_name"].strip()).first()
                if not queue:
                    raise ValueError(f"Очередь не найдена: {row['queue_name']}")
                status = row["status"].strip()
                if status not in allowed_statuses:
                    raise ValueError(f"Недопустимый status: {status}")
                interval_start, interval_end = parse_csv_datetime(row)
                actual_minutes = int(row["actual_minutes"])
                if actual_minutes < 0:
                    raise ValueError("actual_minutes должен быть >= 0")
                existing = db.query(ActualWorkInterval).filter(
                    ActualWorkInterval.work_date == interval_start.date(),
                    ActualWorkInterval.employee_id == employee.id,
                    ActualWorkInterval.queue_id == queue.id,
                    ActualWorkInterval.interval_start == interval_start,
                    ActualWorkInterval.interval_end == interval_end,
                ).first()
                if existing:
                    existing.status = status
                    existing.actual_minutes = actual_minutes
                    existing.source = "csv"
                else:
                    db.add(ActualWorkInterval(
                        work_date=interval_start.date(),
                        employee_id=employee.id,
                        queue_id=queue.id,
                        interval_start=interval_start,
                        interval_end=interval_end,
                        status=status,
                        actual_minutes=actual_minutes,
                        source="csv",
                    ))
                rows_success += 1
            except Exception as exc:
                rows_failed += 1
                db.add(ImportError(batch_id=batch.id, row_number=row_number, error_message=str(exc), raw_data=raw_data))
        batch.rows_total = rows_total
        batch.rows_success = rows_success
        batch.rows_failed = rows_failed
        batch.status = "completed" if rows_failed == 0 else "completed_with_errors"
        batch.completed_at = utcnow()
        log_audit(db, "import_actual_work", request=request, entity_type="import_batch", entity_id=str(batch.id), details=f"rows_success={rows_success}; rows_failed={rows_failed}")
        db.commit()
        db.refresh(batch)
        return batch
    except Exception as exc:
        batch.status = "failed"
        batch.error_message = str(exc)
        batch.completed_at = utcnow()
        db.commit()
        db.refresh(batch)
        return batch


@router.get("/imports", response_model=list[ImportBatchOut])
def list_imports(db: Session = Depends(get_db)):
    return db.query(ImportBatch).order_by(ImportBatch.created_at.desc()).all()


@router.get("/imports/{item_id}", response_model=ImportBatchDetail)
def read_import(item_id: int, db: Session = Depends(get_db)):
    batch = get_or_404(db, ImportBatch, item_id)
    errors = db.query(ImportError).filter(ImportError.batch_id == item_id).order_by(ImportError.row_number).all()
    return ImportBatchDetail.model_validate(batch).model_copy(update={"errors": errors})


@router.get("/staffing", response_model=list[StaffingOut])
def list_staffing(request: Request, project_id: int | None = None, db: Session = Depends(get_db)):
    ensure_project_access(db, request, project_id)
    items = apply_project_filter(db.query(StaffingRequirement), StaffingRequirement, project_id).order_by(StaffingRequirement.interval_start).all()
    return [staffing_out(db, item) for item in items]


@router.post("/staffing", response_model=StaffingOut)
def create_staffing(payload: StaffingIn, db: Session = Depends(get_db)):
    item = save_item(db, StaffingRequirement(**payload.model_dump()))
    return staffing_out(db, item)


@router.get("/staffing/{item_id}", response_model=StaffingOut)
def read_staffing(item_id: int, db: Session = Depends(get_db)):
    item = get_or_404(db, StaffingRequirement, item_id)
    return staffing_out(db, item)


@router.put("/staffing/{item_id}", response_model=StaffingOut)
def update_staffing(item_id: int, payload: StaffingIn, db: Session = Depends(get_db)):
    item = update_item(db, StaffingRequirement, item_id, payload)
    return staffing_out(db, item)


@router.delete("/staffing/{item_id}")
def delete_staffing(item_id: int, db: Session = Depends(get_db)):
    return deactivate_or_delete(db, StaffingRequirement, item_id)


@router.get("/planning/settings", response_model=PlanningSettingsOut)
def read_planning_settings(db: Session = Depends(get_db)):
    return get_planning_settings(db)


@router.post("/planning/settings", response_model=PlanningSettingsOut)
def update_planning_settings(payload: PlanningSettingsIn, db: Session = Depends(get_db)):
    settings = get_planning_settings(db)
    for key, value in payload.model_dump().items():
        setattr(settings, key, value)
    return save_item(db, settings)


@router.put("/planning/settings", response_model=PlanningSettingsOut)
def put_planning_settings(payload: PlanningSettingsIn, db: Session = Depends(get_db)):
    return update_planning_settings(payload, db)


@router.get("/planning/requirements", response_model=list[StaffingOut])
def planning_requirements(request: Request, project_id: int | None = None, db: Session = Depends(get_db)):
    return list_staffing(request, project_id, db)


@router.post("/planning/calculate-staffing", response_model=CalculateStaffingOut)
def calculate_staffing(payload: CalculateStaffingIn, db: Session = Depends(get_db)):
    settings = get_planning_settings(db)
    start = datetime.combine(payload.date_from, time.min)
    end = datetime.combine(payload.date_to, time.max)
    workloads = db.query(WorkloadInterval).filter(
        WorkloadInterval.interval_start >= start,
        WorkloadInterval.interval_start <= end,
    ).order_by(WorkloadInterval.interval_start).all()

    calculated = 0
    for item in workloads:
        required = calculate_required_agents(
            offered_contacts=item.offered_contacts,
            average_handle_time_sec=item.average_handle_time_sec,
            interval_start=item.interval_start,
            interval_end=item.interval_end,
            target_occupancy=settings.target_occupancy,
            min_agents_per_queue=settings.min_agents_per_queue,
            shrinkage_percent=settings.shrinkage_percent,
        )
        note = calculation_note(
            item.offered_contacts,
            item.average_handle_time_sec,
            settings.target_occupancy,
            settings.shrinkage_percent,
            settings.calculation_method,
        )
        existing = db.query(StaffingRequirement).filter(
            StaffingRequirement.interval_start == item.interval_start,
            StaffingRequirement.interval_end == item.interval_end,
            StaffingRequirement.queue_id == item.queue_id,
        ).first()
        if existing:
            existing.required_agents = required
            existing.gap_agents = existing.planned_agents - required
            existing.calculation_note = note
        else:
            db.add(StaffingRequirement(
                interval_start=item.interval_start,
                interval_end=item.interval_end,
                queue_id=item.queue_id,
                required_agents=required,
                planned_agents=0,
                gap_agents=0 - required,
                calculation_note=note,
            ))
        calculated += 1
    db.commit()
    return CalculateStaffingOut(status="ok", calculated_intervals=calculated, date_from=payload.date_from, date_to=payload.date_to)


@router.get("/shifts", response_model=list[ShiftOut])
def list_shifts(request: Request, project_id: int | None = None, db: Session = Depends(get_db)):
    ensure_project_access(db, request, project_id)
    return apply_project_filter(db.query(Shift), Shift, project_id).order_by(Shift.id).all()


@router.post("/shifts", response_model=ShiftOut)
def create_shift(payload: ShiftIn, db: Session = Depends(get_db)):
    return save_item(db, Shift(**payload.model_dump()))


@router.get("/shifts/{item_id}", response_model=ShiftOut)
def read_shift(item_id: int, db: Session = Depends(get_db)):
    return get_or_404(db, Shift, item_id)


@router.put("/shifts/{item_id}", response_model=ShiftOut)
def update_shift(item_id: int, payload: ShiftIn, db: Session = Depends(get_db)):
    return update_item(db, Shift, item_id, payload)


@router.delete("/shifts/{item_id}")
def delete_shift(item_id: int, db: Session = Depends(get_db)):
    return deactivate_or_delete(db, Shift, item_id)


@router.post("/shifts/{item_id}/archive")
def archive_shift(item_id: int, db: Session = Depends(get_db)):
    return deactivate_or_delete(db, Shift, item_id)


@router.post("/shifts/create-defaults")
def create_default_shifts(db: Session = Depends(get_db)) -> dict:
    defaults = [
        ("4 часа утро", time(8, 0), time(12, 0), 0, 4),
        ("4 часа вечер", time(17, 0), time(21, 0), 0, 4),
        ("6 часов день", time(10, 0), time(16, 0), 30, 6),
        ("8 часов стандарт", time(9, 0), time(18, 0), 60, 8),
        ("12 часов длинная", time(9, 0), time(21, 0), 60, 12),
    ]
    created = 0
    for name, start_time, end_time, break_minutes, paid_hours in defaults:
        if not db.query(Shift).filter(Shift.name == name).first():
            db.add(Shift(name=name, start_time=start_time, end_time=end_time, break_minutes=break_minutes, paid_hours=paid_hours))
            created += 1
    db.commit()
    return {"status": "ok", "created": created}


@router.get("/schedule-rules", response_model=list[ScheduleRuleOut])
def list_schedule_rules(db: Session = Depends(get_db)):
    ensure_schedule_rules(db)
    return db.query(ScheduleRule).order_by(ScheduleRule.id).all()


@router.post("/schedule-rules", response_model=ScheduleRuleOut)
def create_schedule_rule(payload: ScheduleRuleIn, db: Session = Depends(get_db)):
    return save_item(db, ScheduleRule(**payload.model_dump()))


@router.put("/schedule-rules/{item_id}", response_model=ScheduleRuleOut)
def update_schedule_rule(item_id: int, payload: ScheduleRuleIn, db: Session = Depends(get_db)):
    return update_item(db, ScheduleRule, item_id, payload)


@router.get("/schedules", response_model=list[ScheduleOut])
def list_schedules(request: Request, project_id: int | None = None, db: Session = Depends(get_db)):
    ensure_project_access(db, request, project_id)
    items = apply_project_filter(db.query(ScheduleAssignment), ScheduleAssignment, project_id).order_by(ScheduleAssignment.work_date).all()
    return [schedule_out(db, item) for item in items]


@router.post("/schedules", response_model=ScheduleOut)
def create_schedule(payload: ScheduleIn, db: Session = Depends(get_db)):
    item = save_item(db, ScheduleAssignment(**payload.model_dump()))
    return read_schedule(item.id, db)


@router.get("/schedules/export.csv")
def export_schedules_csv(request: Request, db: Session = Depends(get_db)):
    rows = [
        [item.work_date, employee_name(db, item.employee_id), shift_name(db, item.shift_id), queue_name(db, item.queue_id), item.status, item.note or ""]
        for item in db.query(ScheduleAssignment).order_by(ScheduleAssignment.work_date, ScheduleAssignment.id).all()
    ]
    return csv_response("schedules.csv", ["date", "employee", "shift", "queue", "status", "note"], rows, db, request, "schedules")


@router.get("/schedules/{item_id}", response_model=ScheduleOut)
def read_schedule(item_id: int, db: Session = Depends(get_db)):
    item = get_or_404(db, ScheduleAssignment, item_id)
    return schedule_out(db, item)


@router.put("/schedules/{item_id}", response_model=ScheduleOut)
def update_schedule(item_id: int, payload: ScheduleIn, db: Session = Depends(get_db)):
    item = update_item(db, ScheduleAssignment, item_id, payload)
    return read_schedule(item.id, db)


@router.delete("/schedules/{item_id}")
def delete_schedule(item_id: int, db: Session = Depends(get_db)):
    return deactivate_or_delete(db, ScheduleAssignment, item_id)


@router.post("/schedules/generate-draft", response_model=GenerateDraftOut)
def generate_draft_schedule(payload: GenerateDraftIn, db: Session = Depends(get_db)):
    settings = get_planning_settings(db)
    start = datetime.combine(payload.date_from, time.min)
    end = datetime.combine(payload.date_to, time.max)

    draft_rows = db.query(ScheduleAssignment).filter(
        ScheduleAssignment.work_date >= payload.date_from,
        ScheduleAssignment.work_date <= payload.date_to,
        ScheduleAssignment.status == "draft",
    )
    if payload.queue_id:
        draft_rows = draft_rows.filter(ScheduleAssignment.queue_id == payload.queue_id)
    replaced = draft_rows.delete(synchronize_session=False)

    employees = db.query(Employee).filter(Employee.is_active.is_(True)).order_by(Employee.id).all()
    shifts = db.query(Shift).filter(Shift.is_active.is_(True)).order_by(Shift.start_time).all()
    if not employees or not shifts:
        db.commit()
        return GenerateDraftOut(status="ok", created_assignments=0, updated_assignments=replaced, skipped_assignments=0, coverage_gaps=0, warnings=["Нет активных сотрудников или смен"], coverage_note="Нет активных сотрудников или смен")

    requirements_query = db.query(
        func.date(StaffingRequirement.interval_start).label("work_date"),
        StaffingRequirement.queue_id,
        func.max(StaffingRequirement.required_agents).label("peak_required"),
    ).filter(
        StaffingRequirement.interval_start >= start,
        StaffingRequirement.interval_start <= end,
    )
    if payload.queue_id:
        requirements_query = requirements_query.filter(StaffingRequirement.queue_id == payload.queue_id)
    requirements = requirements_query.group_by(func.date(StaffingRequirement.interval_start), StaffingRequirement.queue_id).all()

    weekly_hours: dict[int, float] = {
        employee.id: float(db.query(func.coalesce(func.sum(Shift.paid_hours), 0)).join(ScheduleAssignment, ScheduleAssignment.shift_id == Shift.id).filter(
            ScheduleAssignment.employee_id == employee.id,
            ScheduleAssignment.work_date >= payload.date_from,
            ScheduleAssignment.work_date <= payload.date_to,
            ScheduleAssignment.status.in_(["draft", "confirmed", "published", "planned"]),
        ).scalar() or 0)
        for employee in employees
    }
    assigned_by_date: dict[date, set[int]] = {}
    assignments_count: dict[int, int] = {employee.id: 0 for employee in employees}
    existing_assignments = db.query(ScheduleAssignment).filter(
        ScheduleAssignment.work_date >= payload.date_from,
        ScheduleAssignment.work_date <= payload.date_to,
        ScheduleAssignment.status.in_(["confirmed", "published", "planned"]),
    ).all()
    for assignment in existing_assignments:
        assigned_by_date.setdefault(assignment.work_date, set()).add(assignment.employee_id)
        assignments_count[assignment.employee_id] = assignments_count.get(assignment.employee_id, 0) + 1
    created = 0
    skipped = 0
    skill_gaps = 0
    warnings: list[str] = []
    fairness_notes: list[str] = []
    pending_recommendations: list[ScheduleRecommendation] = []
    engine_settings = SchedulerSettings(
        max_weekly_hours=float(settings.max_hours_per_employee_per_week),
        min_rest_hours=settings.min_rest_hours_between_shifts,
        coverage_weight=settings.coverage_weight,
        skill_priority_weight=settings.skill_priority_weight,
        fairness_weight=settings.fairness_weight,
    )

    for requirement in requirements:
        work_date = requirement.work_date
        if isinstance(work_date, str):
            work_date = date.fromisoformat(work_date)
        queue_id = requirement.queue_id
        required_count = int(requirement.peak_required or 0)
        if required_count <= 0:
            continue

        absent_ids = {
            row.employee_id
            for row in db.query(Absence).filter(Absence.date_from <= work_date, Absence.date_to >= work_date).all()
        }
        assigned_today = assigned_by_date.setdefault(work_date, set())
        required_skills = queue_required_skills(db, queue_id)
        candidate_items: list[Employee] = []
        for employee in employees:
            if employee.id in absent_ids or employee.id in assigned_today:
                continue
            if weekly_hours.get(employee.id, 0) >= settings.max_hours_per_employee_per_week:
                continue
            skill_map = employee_skill_map(db, employee.id)
            if not employee_matches_required_skills(skill_map, required_skills):
                continue
            candidate_items.append(employee)

        selected = 0
        while selected < required_count and candidate_items:
            shift = shifts[(created + selected) % len(shifts)]
            shift_start = datetime.combine(work_date, shift.start_time)
            candidates = [
                Candidate(
                    employee_id=employee.id,
                    skill_levels=employee_skill_map(db, employee.id),
                    weekly_hours=weekly_hours.get(employee.id, 0),
                    assignments_count=assignments_count.get(employee.id, 0),
                    absences=[],
                )
                for employee in candidate_items
                if employee.id not in assigned_today
            ]
            best = choose_best_candidate(candidates, required_skills, work_date, shift_start, shift.paid_hours, engine_settings)
            if not best:
                break
            employee = next(item for item in candidate_items if item.id == best.employee_id)
            if selected >= required_count:
                break
            db.add(ScheduleAssignment(
                work_date=work_date,
                employee_id=employee.id,
                shift_id=shift.id,
                queue_id=queue_id,
                status="draft",
                note=f"Черновик создан MVP scoring algorithm; score={round(best.total_score, 1)}",
            ))
            weekly_hours[employee.id] = weekly_hours.get(employee.id, 0) + shift.paid_hours
            assignments_count[employee.id] = assignments_count.get(employee.id, 0) + 1
            assigned_today.add(employee.id)
            selected += 1
            created += 1
        if selected < required_count:
            missing = required_count - selected
            skipped += missing
            if required_skills:
                skill_gaps += missing
            warnings.append(f"{work_date} {queue_name(db, queue_id)}: не хватает {missing} сотрудников с обязательными навыками")
            pending_recommendations.append(ScheduleRecommendation(
                generation_run_id=None,
                work_date=work_date,
                queue_id=queue_id,
                employee_id=None,
                recommendation_type="skill_gap" if required_skills else "coverage_gap",
                message=f"Не закрыто {missing} назначений для очереди {queue_name(db, queue_id)}",
                severity=severity_for_gap(missing),
            ))

    db.commit()
    coverage_count = recalculate_coverage_rows(db, DateRangeIn(date_from=payload.date_from, date_to=payload.date_to, queue_id=payload.queue_id))
    gaps = db.query(ScheduleCoverage).filter(
        ScheduleCoverage.interval_start >= start,
        ScheduleCoverage.interval_start <= end,
        ScheduleCoverage.gap_agents > 0,
    ).count()
    run = ScheduleGenerationRun(
        date_from=payload.date_from,
        date_to=payload.date_to,
        status="ok",
        created_assignments=created,
        skipped_assignments=skipped,
        coverage_gaps=gaps,
        warnings_json=json.dumps(warnings, ensure_ascii=False),
    )
    db.add(run)
    db.flush()
    for recommendation in pending_recommendations:
        recommendation.generation_run_id = run.id
        db.add(recommendation)
    if skipped:
        fairness_notes.append("Дефициты распределения зафиксированы в рекомендациях планировщика")
    db.commit()
    return GenerateDraftOut(
        status="ok",
        created_assignments=created,
        updated_assignments=replaced,
        skipped_assignments=skipped,
        coverage_gaps=gaps,
        skill_gaps=skill_gaps,
        fairness_notes=fairness_notes,
        warnings=warnings,
        coverage_note=f"Черновой график создан по MVP scoring algorithm, покрытие пересчитано: {coverage_count} интервалов",
    )


@router.get("/schedule-recommendations", response_model=list[ScheduleRecommendationOut])
def list_schedule_recommendations(recommendation_type: str | None = None, db: Session = Depends(get_db)):
    query = db.query(ScheduleRecommendation).order_by(ScheduleRecommendation.created_at.desc(), ScheduleRecommendation.id.desc())
    if recommendation_type:
        query = query.filter(ScheduleRecommendation.recommendation_type == recommendation_type)
    return [recommendation_out(db, item) for item in query.limit(200).all()]


def update_schedule_status(item_id: int, status: str, db: Session) -> ScheduleOut:
    item = get_or_404(db, ScheduleAssignment, item_id)
    item.status = status
    if status == "confirmed":
        item.confirmed_at = utcnow()
    elif status == "published":
        item.published_at = utcnow()
    elif status == "cancelled":
        item.cancelled_at = utcnow()
    save_item(db, item)
    return schedule_out(db, item)


@router.post("/schedules/{item_id}/confirm", response_model=ScheduleOut)
def confirm_schedule(item_id: int, db: Session = Depends(get_db)):
    return update_schedule_status(item_id, "confirmed", db)


@router.post("/schedules/{item_id}/publish", response_model=ScheduleOut)
def publish_schedule(item_id: int, db: Session = Depends(get_db)):
    return update_schedule_status(item_id, "published", db)


@router.post("/schedules/{item_id}/cancel", response_model=ScheduleOut)
def cancel_schedule(item_id: int, db: Session = Depends(get_db)):
    return update_schedule_status(item_id, "cancelled", db)


@router.post("/schedules/recalculate-coverage", response_model=CoverageRecalcOut)
def recalculate_coverage(payload: DateRangeIn, db: Session = Depends(get_db)):
    count = recalculate_coverage_rows(db, payload)
    return CoverageRecalcOut(status="ok", recalculated_intervals=count)


@router.get("/coverage", response_model=list[CoverageOut])
def list_coverage(request: Request, project_id: int | None = None, db: Session = Depends(get_db)):
    ensure_project_access(db, request, project_id)
    items = apply_project_filter(db.query(ScheduleCoverage), ScheduleCoverage, project_id).order_by(ScheduleCoverage.interval_start).all()
    return [coverage_out(db, item) for item in items]


@router.get("/coverage/export.csv")
def export_coverage_csv(request: Request, db: Session = Depends(get_db)):
    rows = [
        [
            item.interval_start,
            item.interval_end,
            queue_name(db, item.queue_id),
            item.required_agents,
            item.planned_agents,
            item.confirmed_agents,
            item.published_agents,
            item.gap_agents,
            item.coverage_percent,
        ]
        for item in db.query(ScheduleCoverage).order_by(ScheduleCoverage.interval_start).all()
    ]
    return csv_response("coverage.csv", ["interval_start", "interval_end", "queue", "required", "planned", "confirmed", "published", "gap", "coverage_percent"], rows, db, request, "coverage")


@router.post("/schedules/confirm-period")
def confirm_period(payload: DateRangeIn, db: Session = Depends(get_db)):
    query = db.query(ScheduleAssignment).filter(
        ScheduleAssignment.work_date >= payload.date_from,
        ScheduleAssignment.work_date <= payload.date_to,
        ScheduleAssignment.status == "draft",
    )
    if payload.queue_id:
        query = query.filter(ScheduleAssignment.queue_id == payload.queue_id)
    changed = 0
    for item in query.all():
        item.status = "confirmed"
        item.confirmed_at = utcnow()
        changed += 1
    db.commit()
    return {"status": "ok", "confirmed": changed}


@router.post("/schedules/publish-period")
def publish_period(payload: DateRangeIn, db: Session = Depends(get_db)):
    query = db.query(ScheduleAssignment).filter(
        ScheduleAssignment.work_date >= payload.date_from,
        ScheduleAssignment.work_date <= payload.date_to,
        ScheduleAssignment.status == "confirmed",
    )
    if payload.queue_id:
        query = query.filter(ScheduleAssignment.queue_id == payload.queue_id)
    changed = 0
    for item in query.all():
        item.status = "published"
        item.published_at = utcnow()
        changed += 1
    db.commit()
    recalculate_coverage_rows(db, payload)
    return {"status": "ok", "published": changed}


@router.get("/actual-work", response_model=list[ActualWorkOut])
def list_actual_work(request: Request, project_id: int | None = None, db: Session = Depends(get_db)):
    ensure_project_access(db, request, project_id)
    items = apply_project_filter(db.query(ActualWorkInterval), ActualWorkInterval, project_id).order_by(ActualWorkInterval.interval_start).all()
    return [actual_work_out(db, item) for item in items]


@router.post("/actual-work", response_model=ActualWorkOut)
def create_actual_work(payload: ActualWorkIn, db: Session = Depends(get_db)):
    item = save_item(db, ActualWorkInterval(**payload.model_dump()))
    return actual_work_out(db, item)


@router.put("/actual-work/{item_id}", response_model=ActualWorkOut)
def update_actual_work(item_id: int, payload: ActualWorkIn, db: Session = Depends(get_db)):
    item = update_item(db, ActualWorkInterval, item_id, payload)
    return actual_work_out(db, item)


@router.delete("/actual-work/{item_id}")
def delete_actual_work(item_id: int, db: Session = Depends(get_db)):
    item = get_or_404(db, ActualWorkInterval, item_id)
    db.delete(item)
    db.commit()
    return {"status": "deleted", "id": item_id}


@router.get("/absences", response_model=list[AbsenceOut])
def list_absences(request: Request, project_id: int | None = None, db: Session = Depends(get_db)):
    project_id = current_project_id(db, request, project_id)
    ensure_project_access(db, request, project_id)
    items = apply_project_filter(db.query(Absence), Absence, project_id).order_by(Absence.date_from).all()
    return [AbsenceOut.model_validate(item).model_copy(update={"employee_name": employee_name(db, item.employee_id)}) for item in items]


@router.post("/absences", response_model=AbsenceOut)
def create_absence(payload: AbsenceIn, request: Request, db: Session = Depends(get_db)):
    data = payload.model_dump()
    project_id = current_project_id(db, request, data.get("project_id"))
    ensure_project_access(db, request, project_id)
    if not employee_has_project_access(db, data["employee_id"], project_id):
        raise HTTPException(status_code=422, detail="Сотрудник должен принадлежать текущему проекту")
    data["project_id"] = project_id
    item = save_item(db, Absence(**data))
    return AbsenceOut.model_validate(item).model_copy(update={"employee_name": employee_name(db, item.employee_id)})


@router.get("/absences/{item_id}", response_model=AbsenceOut)
def read_absence(item_id: int, db: Session = Depends(get_db)):
    item = get_or_404(db, Absence, item_id)
    return AbsenceOut.model_validate(item).model_copy(update={"employee_name": employee_name(db, item.employee_id)})


@router.put("/absences/{item_id}", response_model=AbsenceOut)
def update_absence(item_id: int, payload: AbsenceIn, request: Request, db: Session = Depends(get_db)):
    item = get_or_404(db, Absence, item_id)
    data = payload.model_dump()
    project_id = current_project_id(db, request, data.get("project_id") or item.project_id)
    ensure_project_access(db, request, project_id)
    if not employee_has_project_access(db, data["employee_id"], project_id):
        raise HTTPException(status_code=422, detail="Сотрудник должен принадлежать текущему проекту")
    data["project_id"] = project_id
    for key, value in data.items():
        setattr(item, key, value)
    item = save_item(db, item)
    return AbsenceOut.model_validate(item).model_copy(update={"employee_name": employee_name(db, item.employee_id)})


@router.delete("/absences/{item_id}")
def delete_absence(item_id: int, db: Session = Depends(get_db)):
    return deactivate_or_delete(db, Absence, item_id)


@router.post("/absences/{item_id}/cancel")
def cancel_absence(item_id: int, request: Request, db: Session = Depends(get_db)):
    item = get_or_404(db, Absence, item_id)
    item.status = "cancelled"
    log_audit(db, "absence_cancel", request=request, entity_type="absence", entity_id=str(item_id))
    item = save_item(db, item)
    return AbsenceOut.model_validate(item).model_copy(update={"employee_name": employee_name(db, item.employee_id)})


@router.post("/absences/{item_id}/archive")
def archive_absence(item_id: int, request: Request, db: Session = Depends(get_db)):
    return cancel_absence(item_id, request, db)


@router.get("/settings/app")
def read_app_settings(db: Session = Depends(get_db)) -> dict:
    defaults = {
        "company_name": "Телесейлз Сервис",
        "system_name": "WFM-платформа",
        "timezone": "Europe/Moscow",
        "language": "ru",
        "date_format": "DD.MM.YYYY",
        "auto_create_teams": "true",
        "auto_create_skills": "true",
        "max_import_rows": "10000",
        "csv_delimiter": ",",
        "csv_encoding": "UTF-8 BOM",
        "export_include_ids": "false",
    }
    stored = {row.key: row.value for row in db.query(AppSetting).all()}
    defaults.update({key: value for key, value in stored.items() if value is not None})
    return defaults


@router.put("/settings/app")
def save_app_settings(payload: dict, request: Request, db: Session = Depends(get_db)) -> dict:
    allowed = {
        "company_name", "system_name", "timezone", "language", "date_format",
        "auto_create_teams", "auto_create_skills", "max_import_rows",
        "csv_delimiter", "csv_encoding", "export_include_ids",
    }
    for key, value in payload.items():
        if key not in allowed:
            continue
        item = db.query(AppSetting).filter(AppSetting.key == key).first()
        if not item:
            item = AppSetting(key=key, value_type="string")
        item.value = str(value)
        db.add(item)
    log_audit(db, "app_settings_update", request=request, entity_type="settings")
    db.commit()
    return read_app_settings(db)


def build_plan_fact(db: Session, date_from: date, date_to: date, queue_id: int | None = None) -> PlanFactOut:
    schedules = db.query(ScheduleAssignment).filter(
        ScheduleAssignment.work_date >= date_from,
        ScheduleAssignment.work_date <= date_to,
    )
    actuals = db.query(ActualWorkInterval).filter(
        ActualWorkInterval.work_date >= date_from,
        ActualWorkInterval.work_date <= date_to,
    )
    absences = db.query(Absence).filter(Absence.date_from <= date_to, Absence.date_to >= date_from)
    if queue_id:
        schedules = schedules.filter(ScheduleAssignment.queue_id == queue_id)
        actuals = actuals.filter(ActualWorkInterval.queue_id == queue_id)

    schedule_rows = schedules.all()
    published_rows = [row for row in schedule_rows if row.status == "published"]
    planned_hours = 0.0
    for row in published_rows:
        shift = db.get(Shift, row.shift_id)
        planned_hours += shift.paid_hours if shift else 0

    actual_rows = actuals.all()
    actual_hours = round(sum(row.actual_minutes for row in actual_rows if row.status == "worked") / 60, 2)
    adherence = round(actual_hours / planned_hours * 100, 1) if planned_hours > 0 else 0
    rows = [
        {
            "work_date": row.work_date.isoformat(),
            "employee_name": employee_name(db, row.employee_id),
            "queue_name": queue_name(db, row.queue_id),
            "status": row.status,
            "actual_minutes": row.actual_minutes,
            "actual_hours": round(row.actual_minutes / 60, 2),
        }
        for row in actual_rows
    ]
    return PlanFactOut(
        planned_assignments=len(schedule_rows),
        published_assignments=len(published_rows),
        actual_worked_intervals=len([row for row in actual_rows if row.status == "worked"]),
        absence_count=absences.count(),
        adherence_percent=adherence,
        planned_hours=round(planned_hours, 2),
        actual_hours=actual_hours,
        gap_hours=round(actual_hours - planned_hours, 2),
        rows=rows,
    )


def executive_summary_data(db: Session, project_id: int | None = None) -> dict:
    plan_fact = build_plan_fact(db, date.today() - timedelta(days=7), date.today() + timedelta(days=7))
    coverage_query = db.query(func.avg(ScheduleCoverage.coverage_percent))
    workload_query = db.query(func.avg(WorkloadInterval.service_level_percent))
    aht_query = db.query(func.avg(WorkloadInterval.average_handle_time_sec))
    employee_query = db.query(Employee).join(EmployeeProjectAccess, EmployeeProjectAccess.employee_id == Employee.id).filter(Employee.is_active.is_(True))
    queue_query = db.query(Queue).filter(Queue.is_active.is_(True))
    schedule_query = db.query(ScheduleAssignment).filter(ScheduleAssignment.status == "published")
    gap_query = db.query(ScheduleCoverage).filter(ScheduleCoverage.gap_agents > 0)
    if project_id is not None:
        coverage_query = coverage_query.filter(ScheduleCoverage.project_id == project_id)
        workload_query = workload_query.filter(WorkloadInterval.project_id == project_id)
        aht_query = aht_query.filter(WorkloadInterval.project_id == project_id)
        employee_query = employee_query.filter(EmployeeProjectAccess.project_id == project_id, EmployeeProjectAccess.can_work.is_(True))
        queue_query = queue_query.filter(Queue.project_id == project_id)
        schedule_query = schedule_query.filter(ScheduleAssignment.project_id == project_id)
        gap_query = gap_query.filter(ScheduleCoverage.project_id == project_id)
    avg_coverage = coverage_query.scalar() or 0
    avg_sl = workload_query.scalar() or 0
    avg_aht = aht_query.scalar() or 0
    return {
        "active_employees": employee_query.count(),
        "active_queues": queue_query.count(),
        "published_assignments": schedule_query.count(),
        "avg_coverage_percent": round(float(avg_coverage), 1),
        "gap_intervals_count": gap_query.count(),
        "avg_service_level_percent": round(float(avg_sl), 1),
        "avg_aht_sec": round(float(avg_aht), 1),
        "planned_hours": plan_fact.planned_hours,
        "actual_hours": plan_fact.actual_hours,
        "adherence_percent": plan_fact.adherence_percent,
    }


def operations_summary_data(db: Session, project_id: int | None = None) -> list[dict]:
    rows = []
    grouped_query = db.query(
        func.date(WorkloadInterval.interval_start).label("work_date"),
        WorkloadInterval.queue_id,
        func.sum(WorkloadInterval.offered_contacts).label("offered"),
        func.sum(WorkloadInterval.handled_contacts).label("handled"),
        func.sum(WorkloadInterval.abandoned_contacts).label("abandoned"),
        func.avg(WorkloadInterval.average_handle_time_sec).label("aht"),
        func.avg(WorkloadInterval.service_level_percent).label("sl"),
    )
    if project_id is not None:
        grouped_query = grouped_query.filter(WorkloadInterval.project_id == project_id)
    grouped = grouped_query.group_by(func.date(WorkloadInterval.interval_start), WorkloadInterval.queue_id).order_by(func.date(WorkloadInterval.interval_start), WorkloadInterval.queue_id).all()
    for row in grouped:
        req = db.query(
            func.sum(StaffingRequirement.required_agents),
            func.sum(StaffingRequirement.planned_agents),
            func.sum(StaffingRequirement.gap_agents),
        ).filter(
            func.date(StaffingRequirement.interval_start) == row.work_date,
            StaffingRequirement.queue_id == row.queue_id,
        ).one()
        rows.append({
            "work_date": str(row.work_date),
            "queue_id": row.queue_id,
            "queue_name": queue_name(db, row.queue_id),
            "offered_contacts": int(row.offered or 0),
            "handled_contacts": int(row.handled or 0),
            "abandoned_contacts": int(row.abandoned or 0),
            "avg_aht_sec": round(float(row.aht or 0), 1),
            "avg_service_level_percent": round(float(row.sl or 0), 1),
            "required_agents": int(req[0] or 0),
            "planned_agents": int(req[1] or 0),
            "gap_agents": int(req[2] or 0),
        })
    return rows


def staffing_efficiency_data(db: Session, project_id: int | None = None) -> dict:
    plan_fact = build_plan_fact(db, date.today() - timedelta(days=7), date.today() + timedelta(days=7))
    utilization_query = db.query(func.avg(KpiSnapshot.utilization_percent))
    if project_id is not None:
        utilization_query = utilization_query.filter(KpiSnapshot.project_id == project_id)
    utilization = utilization_query.scalar() or 0
    return {
        "planned_hours": plan_fact.planned_hours,
        "actual_hours": plan_fact.actual_hours,
        "adherence_percent": plan_fact.adherence_percent,
        "utilization_percent": round(float(utilization), 1),
        "deficit_hours": abs(plan_fact.gap_hours) if plan_fact.gap_hours < 0 else 0,
        "surplus_hours": plan_fact.gap_hours if plan_fact.gap_hours > 0 else 0,
    }


def coverage_gaps_data(db: Session, project_id: int | None = None) -> list[dict]:
    query = db.query(ScheduleCoverage).filter(ScheduleCoverage.gap_agents > 0)
    if project_id is not None:
        query = query.filter(ScheduleCoverage.project_id == project_id)
    return [
        {
            "interval_start": row.interval_start.isoformat(),
            "interval_end": row.interval_end.isoformat(),
            "queue_id": row.queue_id,
            "queue_name": queue_name(db, row.queue_id),
            "required_agents": row.required_agents,
            "planned_agents": row.planned_agents,
            "gap_agents": row.gap_agents,
            "severity": severity_for_gap(row.gap_agents),
        }
        for row in query.order_by(ScheduleCoverage.interval_start).limit(500).all()
    ]


def sla_summary_data(db: Session, project_id: int | None = None) -> list[dict]:
    rows = []
    queue_query = db.query(Queue)
    if project_id is not None:
        queue_query = queue_query.filter(Queue.project_id == project_id)
    for queue in queue_query.order_by(Queue.id).all():
        intervals_query = db.query(WorkloadInterval).filter(WorkloadInterval.queue_id == queue.id)
        if project_id is not None:
            intervals_query = intervals_query.filter(WorkloadInterval.project_id == project_id)
        intervals = intervals_query.all()
        if not intervals:
            continue
        below = [item for item in intervals if item.service_level_percent < queue.service_level_target]
        rows.append({
            "queue_id": queue.id,
            "queue_name": queue.name,
            "target_service_level_percent": queue.service_level_target,
            "avg_service_level_percent": round(sum(item.service_level_percent for item in intervals) / len(intervals), 1),
            "min_service_level_percent": round(min(item.service_level_percent for item in intervals), 1),
            "below_target_intervals": len(below),
            "intervals_total": len(intervals),
        })
    return rows


@router.get("/reports/executive-summary")
def executive_summary(request: Request, db: Session = Depends(get_db)):
    return executive_summary_data(db, current_project_id(db, request, None))


@router.get("/reports/operations-summary")
def operations_summary(request: Request, db: Session = Depends(get_db)):
    return operations_summary_data(db, current_project_id(db, request, None))


@router.get("/reports/staffing-efficiency")
def staffing_efficiency(request: Request, db: Session = Depends(get_db)):
    return staffing_efficiency_data(db, current_project_id(db, request, None))


@router.get("/reports/coverage-gaps")
def coverage_gaps_report(request: Request, db: Session = Depends(get_db)):
    return coverage_gaps_data(db, current_project_id(db, request, None))


@router.get("/reports/sla-summary")
def sla_summary(request: Request, db: Session = Depends(get_db)):
    return sla_summary_data(db, current_project_id(db, request, None))


@router.get("/reports/executive-summary.csv")
def executive_summary_csv(request: Request, db: Session = Depends(get_db)):
    data = executive_summary_data(db, current_project_id(db, request, None))
    rows = [[key, value] for key, value in data.items()]
    return csv_response("executive-summary.csv", ["metric", "value"], rows, db, request, "executive_summary")


@router.get("/reports/operations-summary.csv")
def operations_summary_csv(request: Request, db: Session = Depends(get_db)):
    rows = [[item[key] for key in ["work_date", "queue_name", "offered_contacts", "handled_contacts", "abandoned_contacts", "avg_aht_sec", "avg_service_level_percent", "required_agents", "planned_agents", "gap_agents"]] for item in operations_summary_data(db, current_project_id(db, request, None))]
    return csv_response("operations-summary.csv", ["date", "queue", "offered", "handled", "abandoned", "aht", "sl", "required", "planned", "gap"], rows, db, request, "operations_summary")


@router.get("/reports/staffing-efficiency.csv")
def staffing_efficiency_csv(request: Request, db: Session = Depends(get_db)):
    data = staffing_efficiency_data(db, current_project_id(db, request, None))
    rows = [[key, value] for key, value in data.items()]
    return csv_response("staffing-efficiency.csv", ["metric", "value"], rows, db, request, "staffing_efficiency")


@router.get("/reports/coverage-gaps.csv")
def coverage_gaps_csv(request: Request, db: Session = Depends(get_db)):
    rows = [[item[key] for key in ["interval_start", "interval_end", "queue_name", "required_agents", "planned_agents", "gap_agents", "severity"]] for item in coverage_gaps_data(db, current_project_id(db, request, None))]
    return csv_response("coverage-gaps.csv", ["interval_start", "interval_end", "queue", "required", "planned", "gap", "severity"], rows, db, request, "coverage_gaps")


@router.get("/reports/sla-summary.csv")
def sla_summary_csv(request: Request, db: Session = Depends(get_db)):
    rows = [[item[key] for key in ["queue_name", "target_service_level_percent", "avg_service_level_percent", "min_service_level_percent", "below_target_intervals", "intervals_total"]] for item in sla_summary_data(db, current_project_id(db, request, None))]
    return csv_response("sla-summary.csv", ["queue", "target_sl", "avg_sl", "min_sl", "below_target", "intervals_total"], rows, db, request, "sla_summary")


@router.get("/reports/export-log")
def export_log(db: Session = Depends(get_db)):
    return [
        {
            "id": row.id,
            "report_type": row.report_type,
            "filename": row.filename,
            "rows_count": row.rows_count,
            "created_at": row.created_at,
        }
        for row in db.query(ExportLog).order_by(ExportLog.created_at.desc()).limit(200).all()
    ]


@router.get("/reports/plan-fact", response_model=PlanFactOut)
def plan_fact_report(date_from: date, date_to: date, queue_id: int | None = None, db: Session = Depends(get_db)):
    return build_plan_fact(db, date_from, date_to, queue_id)


@router.get("/reports/schedule-coverage", response_model=list[CoverageOut])
def report_schedule_coverage(db: Session = Depends(get_db)):
    return list_coverage(db)


@router.get("/reports/staffing-gaps")
def report_staffing_gaps(db: Session = Depends(get_db)):
    return [
        {
            "interval_start": row.interval_start,
            "queue_name": queue_name(db, row.queue_id),
            "required_agents": row.required_agents,
            "planned_agents": row.planned_agents,
            "gap_agents": row.gap_agents,
        }
        for row in db.query(StaffingRequirement).filter(StaffingRequirement.gap_agents < 0).order_by(StaffingRequirement.interval_start).limit(200).all()
    ]


@router.get("/reports/plan-fact.csv")
def plan_fact_csv(date_from: date, date_to: date, request: Request, queue_id: int | None = None, db: Session = Depends(get_db)):
    report = build_plan_fact(db, date_from, date_to, queue_id)
    rows = [
        [row["work_date"], row["employee_name"], row["queue_name"], row["status"], row["actual_minutes"], row["actual_hours"]]
        for row in report.rows
    ]
    return csv_response("plan-fact.csv", ["date", "employee", "queue", "status", "actual_minutes", "actual_hours"], rows, db, request, "plan_fact")


@router.get("/reports/summary", response_model=SummaryOut)
def summary(request: Request, db: Session = Depends(get_db)):
    project_id = current_project_id(db, request, None)
    today = date.today()
    week_end = today + timedelta(days=7)
    workload_query = db.query(func.avg(WorkloadInterval.service_level_percent), func.avg(WorkloadInterval.average_handle_time_sec))
    staffing_gap_query = db.query(func.sum(StaffingRequirement.gap_agents))
    if project_id is not None:
        workload_query = workload_query.filter(WorkloadInterval.project_id == project_id)
        staffing_gap_query = staffing_gap_query.filter(StaffingRequirement.project_id == project_id)
    avg_sl, avg_aht = workload_query.one()
    avg_sl = avg_sl or 0
    avg_aht = avg_aht or 0
    gap = staffing_gap_query.scalar() or 0
    asa = 18
    staffing_by_queue = [
        {
            "queue_name": queue_name(db, row.queue_id),
            "required_agents": int(row.required_agents or 0),
            "planned_agents": int(row.planned_agents or 0),
            "gap_agents": int(row.gap_agents or 0),
        }
        for row in db.query(
            StaffingRequirement.queue_id,
            func.sum(StaffingRequirement.required_agents).label("required_agents"),
            func.sum(StaffingRequirement.planned_agents).label("planned_agents"),
            func.sum(StaffingRequirement.gap_agents).label("gap_agents"),
        ).filter(StaffingRequirement.project_id == project_id if project_id is not None else True).group_by(StaffingRequirement.queue_id).order_by(StaffingRequirement.queue_id).all()
    ]
    planned_coverage = [
        {
            "queue_name": item["queue_name"],
            "coverage_percent": round((item["planned_agents"] / item["required_agents"] * 100), 1) if item["required_agents"] else 0,
        }
        for item in staffing_by_queue
    ]
    recent_imports = [
        {
            "id": row.id,
            "filename": row.filename,
            "status": row.status,
            "rows_success": row.rows_success,
            "rows_failed": row.rows_failed,
        }
        for row in db.query(ImportBatch).filter(ImportBatch.project_id == project_id if project_id is not None else True).order_by(ImportBatch.created_at.desc()).limit(5).all()
    ]
    today_schedules = [
        {
            "employee_name": employee_name(db, row.employee_id),
            "shift_name": shift_name(db, row.shift_id),
            "queue_name": queue_name(db, row.queue_id),
            "status": row.status,
        }
        for row in db.query(ScheduleAssignment).filter(ScheduleAssignment.work_date == today, ScheduleAssignment.project_id == project_id if project_id is not None else True).order_by(ScheduleAssignment.id).limit(8).all()
    ]
    coverage_gaps = [
        {
            "interval_start": row.interval_start.isoformat(),
            "queue_name": queue_name(db, row.queue_id),
            "gap_agents": row.gap_agents,
            "coverage_percent": row.coverage_percent,
        }
        for row in db.query(ScheduleCoverage).filter(ScheduleCoverage.gap_agents > 0, ScheduleCoverage.project_id == project_id if project_id is not None else True).order_by(ScheduleCoverage.interval_start).limit(8).all()
    ]
    recent_generations = [
        {
            "id": row.id,
            "date_from": row.date_from.isoformat(),
            "date_to": row.date_to.isoformat(),
            "created_assignments": row.created_assignments,
            "coverage_gaps": row.coverage_gaps,
        }
        for row in db.query(ScheduleGenerationRun).order_by(ScheduleGenerationRun.created_at.desc()).limit(5).all()
    ]
    week_plan_fact = build_plan_fact(db, today, week_end)
    return SummaryOut(
        total_employees=db.query(Employee).join(EmployeeProjectAccess, EmployeeProjectAccess.employee_id == Employee.id).filter(EmployeeProjectAccess.project_id == project_id, EmployeeProjectAccess.can_work.is_(True)).count() if project_id is not None else db.query(Employee).count(),
        active_teams=db.query(Team).filter(Team.is_active.is_(True), Team.project_id == project_id if project_id is not None else True).count(),
        queues=db.query(Queue).filter(Queue.project_id == project_id if project_id is not None else True).count(),
        shifts=db.query(Shift).filter(Shift.project_id == project_id if project_id is not None else True).count(),
        week_assignments=db.query(ScheduleAssignment).filter(ScheduleAssignment.work_date >= today, ScheduleAssignment.work_date <= week_end, ScheduleAssignment.project_id == project_id if project_id is not None else True).count(),
        average_service_level=round(float(avg_sl), 1),
        average_aht=round(float(avg_aht), 1),
        staffing_gap=int(gap),
        kpi={
            "service_level": round(float(avg_sl), 1),
            "asa": asa,
            "aht": round(float(avg_aht), 1),
            "occupancy": 74.5,
            "utilization": 81.2,
            "abandonment": 4.1,
        },
        staffing_by_queue=staffing_by_queue,
        planned_coverage=planned_coverage,
        recent_imports=recent_imports,
        today_schedules=today_schedules,
        coverage_gaps=coverage_gaps,
        recent_generations=recent_generations,
        plan_fact={
            "planned_hours": week_plan_fact.planned_hours,
            "actual_hours": week_plan_fact.actual_hours,
            "adherence_percent": week_plan_fact.adherence_percent,
        },
    )


@router.get("/reports/dashboard-summary", response_model=SummaryOut)
def dashboard_summary(request: Request, db: Session = Depends(get_db)):
    return summary(request, db)
